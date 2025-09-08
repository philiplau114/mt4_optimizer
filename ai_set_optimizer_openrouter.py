import argparse
import json
import os
import sqlite3
import re
import requests
import logging
import sys
import time
from collections import Counter, defaultdict

from set_file_updater import update_parameters

# --- Logging Setup ---
class FlushFileHandler(logging.FileHandler):
    def emit(self, record):
        super().emit(record)
        self.flush()

LOG_FILE = os.path.join(os.path.dirname(__file__), "ai_set_optimizer_openrouter.log")
log_to_file = False  # Set to False to disable file logging

handlers = []
if log_to_file:
    handlers.append(FlushFileHandler(LOG_FILE, encoding='utf-8'))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=handlers
)
logger = logging.getLogger(__name__)

# --- AI JSON extraction and retry helpers ---
def extract_json_code_blocks(response_text):
    """Extracts all code blocks marked as ```json ... ``` from the response."""
    return re.findall(r"```json(.*?)```", response_text, re.DOTALL)

def parse_json_blocks(blocks):
    """Attempts to parse a list of JSON code blocks. Returns list of valid objects, or empty list."""
    objs = []
    for block in blocks:
        block = block.strip()
        try:
            objs.append(json.loads(block))
        except Exception as e:
            logger.warning(f"Failed to parse JSON block: {e}")
            continue
    return objs

def get_valid_json_from_ai(call_ai_func, prompt, model, max_attempts=3, delay_sec=2, debug_path=None):
    """
    Calls the AI model, checks response for valid JSON code blocks.
    Retries up to max_attempts if parsing fails.
    Raises Exception if all attempts fail.
    Returns: list of valid JSON objects (usually two, per your format).
    """
    last_response = None
    for attempt in range(1, max_attempts + 1):
        logger.info(f"AI JSON parse attempt {attempt} for model {model}")
        response = call_ai_func(prompt, model)
        last_response = response
        blocks = extract_json_code_blocks(response)
        valid_jsons = parse_json_blocks(blocks)
        if len(valid_jsons) == 2:
            logger.info(f"Successfully extracted two valid JSON blocks on attempt {attempt} for model {model}")
            return valid_jsons
        logger.warning(f"Attempt {attempt}: Did not find two valid JSON blocks in AI response. Retrying...")
        time.sleep(delay_sec)
    # All attempts failed
    logger.error("Failed to get two valid JSON blocks from AI after maximum attempts.")
    # Optionally write last_response to a .suggestions.txt file for review
    if debug_path:
        with open(debug_path, "w", encoding="utf-8") as f:
            f.write(last_response)
        logger.info(f"Last failed AI response written to {debug_path}")
    raise Exception("Failed to get two valid JSON blocks from AI after max attempts.")

def read_file_head(path, max_chars=8000):
    try:
        with open(path, encoding='utf-8') as f:
            result = f.read(max_chars)
        logger.info(f"Read {min(len(result), max_chars)} characters from {path}")
        return result
    except Exception as e:
        logger.error(f"Failed to read file head from {path}: {e}")
        return ""

def load_prompt_template(template_path):
    try:
        with open(template_path, encoding="utf-8") as f:
            content = f.read()
        logger.info(f"Loaded prompt template from {template_path}")
        return content
    except Exception as e:
        logger.error(f"Failed to load prompt template from {template_path}: {e}")
        return ""

def build_ancestry_chain(conn, start_set_file_name):
    cur = conn.cursor()
    cur.execute("SELECT set_file_name, input_set_file FROM test_metrics")
    file_map = {row[0]: row[1] for row in cur.fetchall()}
    ancestry = []
    current = start_set_file_name
    while current and current not in ancestry:
        ancestry.append(current)
        input_file = file_map.get(current)
        if input_file:
            parent = os.path.basename(input_file)
            if parent == current:
                break
            current = parent
        else:
            break
    logger.info(f"Ancestry chain for {start_set_file_name}: {ancestry}")
    return ancestry

def parse_param_to_section(spec_content):
    param_to_section = {}
    lines = [l.strip() for l in spec_content.splitlines() if l.strip()]
    if not lines:
        return param_to_section
    header = [x.strip() for x in lines[0].split(",")]
    try:
        section_idx = header.index("Section")
        param_idx = header.index("Parameter")
    except ValueError:
        section_idx = 0
        param_idx = 1
    for line in lines[1:]:
        parts = [x.strip() for x in re.split(r',(?=(?:[^\"]*\"[^\"]*\")*[^\"]*$)', line)]
        if len(parts) > max(section_idx, param_idx):
            section = parts[section_idx]
            pname = parts[param_idx]
            if pname:
                param_to_section[pname] = section
    logger.info(f"Parsed {len(param_to_section)} param-to-section mappings.")
    return param_to_section

def fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section=None):
    if not ancestry:
        return []
    placeholders = ','.join(['?'] * len(ancestry))
    sql = f"""
    SELECT p.parameter_name
    FROM set_file_steps AS s
    JOIN optimization_suggestion AS sug ON s.id = sug.step_id
    JOIN optimization_parameter AS p ON sug.id = p.suggestion_id
    JOIN test_metrics AS tm ON tm.step_id = s.id
    WHERE tm.set_file_name IN ({placeholders})
    """
    cur = conn.cursor()
    cur.execute(sql, ancestry)
    rows = [row[0] for row in cur.fetchall()]
    if param_to_section is not None:
        result = [(param_to_section.get(p, "Unknown Section"), p) for p in rows]
    else:
        result = [("Unknown Section", p) for p in rows]
    logger.info(f"Fetched {len(result)} suggestions for ancestry sections.")
    return result

def make_suggestion_history_summary_block(conn, set_file_name, spec_content=None):
    param_to_section = None
    if spec_content:
        param_to_section = parse_param_to_section(spec_content)
    ancestry = build_ancestry_chain(conn, set_file_name)
    section_param_pairs = fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section)
    section_param_counter = defaultdict(Counter)
    for section, pname in section_param_pairs:
        section_param_counter[section][pname] += 1
    lines = []
    for section in sorted(section_param_counter):
        lines.append(f"Section: {section}")
        for pname, count in sorted(section_param_counter[section].items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {pname}: {count}   # times previously suggested")
        lines.append("")  # Blank line for separation
    logger.info(f"Suggestion history summary block built for {set_file_name}.")
    return "\n".join(lines).strip()

def build_prompt_with_history(
    prompt_template,
    base_parameters,
    set_content,
    phoenix_spec_content,
    summary_csv_content,
    performance_metrics_block,
    suggestion_history_summary_block,
    wave_analysis_result_block
):
    prompt = prompt_template.format(
        base_parameters=", ".join(base_parameters),
        performance_metrics_block=performance_metrics_block,
        suggestion_history_summary_block=suggestion_history_summary_block,
        wave_analysis_result_block=wave_analysis_result_block
    )
    prompt = prompt.replace("[Upload .set FILE or paste content here]", set_content)
    prompt = prompt.replace("[Upload PhoenixSpec.csv or paste content here]", phoenix_spec_content)
    prompt = prompt.replace("[Upload SUMMARY.csv or paste content here]", summary_csv_content)
    logger.info("Prompt with history built.")
    return prompt

def save_optimization_suggestion_to_db(
    db_path, step_id, mode_sections_obj, param_array
):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute(
        "INSERT INTO optimization_suggestion (step_id, mode) VALUES (?, ?)",
        (step_id, mode_sections_obj['mode'])
    )
    suggestion_id = c.lastrowid
    for section in mode_sections_obj['sections']:
        c.execute(
            "INSERT INTO optimization_section (suggestion_id, section_name, explanation) VALUES (?, ?, ?)",
            (suggestion_id, section['name'], section['explanation'])
        )
    for param in param_array:
        c.execute(
            "INSERT INTO optimization_parameter (suggestion_id, parameter_name, start, end, step, reason) VALUES (?, ?, ?, ?, ?, ?)",
            (
                suggestion_id,
                param['name'],
                param.get('start'),
                param.get('end'),
                param.get('step'),
                param.get('reason')
            )
        )
    conn.commit()
    conn.close()
    logger.info(f"Saved optimization suggestion to DB with suggestion_id={suggestion_id}")
    return suggestion_id

import openpyxl

def get_performance_metrics_block(config_xlsx_path, sheet_name="performance_criteria"):
    try:
        wb = openpyxl.load_workbook(config_xlsx_path, data_only=True)
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        lines = []
        for row in rows:
            if row and row[0]:
                key = str(row[0])
                value = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                explanation = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                lines.append(f"{key}: {value}   # {explanation}")
        logger.info(f"Performance metrics block loaded from {config_xlsx_path}.")
        return "\n".join(lines)
    except Exception as e:
        logger.warning(f"Failed to load performance metrics block: {e}")
        return ""

def call_openrouter(prompt, model, api_key):
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are an expert in MetaTrader 4/5 optimization and parameter tuning."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.1,
        "max_tokens": 1600
    }
    try:
        response = requests.post(url, headers=headers, json=data)
        response.raise_for_status()
        result = response.json()
        logger.info(f"OpenRouter API call successful for model: {model}")
        return result["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"OpenRouter API call failed for model {model}: {e}")
        raise

def coverage_voting(all_param_arrays):
    param_ranges = defaultdict(list)
    for param_array in all_param_arrays:
        for s in param_array:
            param_ranges[s["name"]].append((s.get("start"), s.get("end"), s.get("step")))
    merged = []
    for name, ranges in param_ranges.items():
        numeric_ranges = [r for r in ranges if r[0] is not None and r[1] is not None and r[2] is not None]
        if not numeric_ranges:
            continue
        try:
            min_start = min(r[0] for r in numeric_ranges)
            max_end = max(r[1] for r in numeric_ranges)
            min_step = min(r[2] for r in numeric_ranges)
        except Exception as e:
            logger.warning(f"Coverage voting failed for param {name}: {e}")
            continue
        merged.append({
            "name": name,
            "start": min_start,
            "end": max_end,
            "step": min_step,
            "reason": f"Coverage of all model suggestions: {[(r[0], r[1]) for r in ranges]}"
        })
    logger.info(f"Coverage voting merged {len(merged)} parameter ranges.")
    return merged

def suggest_mode_and_sections_and_params_openrouter(
    template_path,
    base_parameters,
    set_path,
    spec_path,
    summary_path,
    openrouter_api_key,
    output_path,
    db_path,
    step_id,
    config_xlsx_path=None,
    suggestion_json_path=None,
    models=None,
    wave_analysis_block=None
):
    import os

    # Defensive: Ensure models is a list, even if passed as a string.
    if models is None:
        models = ["openai/gpt-4o"]
    elif isinstance(models, str):
        models = [m.strip() for m in models.split(",") if m.strip()]

    # Defensive: Check files exist before reading.
    for fpath, label in [
        (template_path, "Prompt template"),
        (set_path, ".set file"),
        (spec_path, "PhoenixSpec.csv"),
        (summary_path, "SUMMARY.csv"),
        (db_path, "Database"),
    ]:
        if not os.path.exists(fpath):
            logger.error(f"{label} does not exist: {fpath}")
            return None

    prompt_template = load_prompt_template(template_path)
    set_content = read_file_head(set_path, 8000)
    phoenix_spec_content = read_file_head(spec_path, 8000)
    summary_csv_content = read_file_head(summary_path, 8000)
    base_parameters_list = [s.strip() for s in base_parameters.split(",") if s.strip()]
    if config_xlsx_path:
        performance_metrics_block = get_performance_metrics_block(config_xlsx_path)
    else:
        performance_metrics_block = ""
    set_file_name = os.path.basename(set_path)
    conn = sqlite3.connect(db_path)
    suggestion_history_summary_block = make_suggestion_history_summary_block(
        conn, set_file_name, spec_content=phoenix_spec_content
    )
    conn.close()
    prompt = build_prompt_with_history(
        prompt_template,
        base_parameters_list,
        set_content,
        phoenix_spec_content,
        summary_csv_content,
        performance_metrics_block,
        suggestion_history_summary_block,
        wave_analysis_block  # <--- NEW: pass here
    )

    logger.info(f"Models: {models}")
    logger.info(f"Prompt size: {len(prompt)} chars")
    try:
        import tiktoken
        enc = tiktoken.encoding_for_model("gpt-4o")
        token_count = len(enc.encode(prompt))
        logger.info(f"Prompt token count: {token_count}")
    except Exception as e:
        logger.warning(f"Could not count prompt tokens: {e}")

    all_mode_sections = []
    all_param_arrays = []
    for model in models:
        logger.info(f"Querying OpenRouter model with retry and JSON validation: {model}")
        try:
            # Use the robust retry logic for valid JSON extraction
            debug_path = output_path + f".{model.replace('/','_')}.suggestions.txt"
            valid_jsons = get_valid_json_from_ai(
                lambda p, m=model: call_openrouter(p, m, openrouter_api_key),
                prompt,
                model=model,
                max_attempts=3,
                delay_sec=2,
                debug_path=debug_path
            )
            mode_sections_obj, param_array = valid_jsons[0], valid_jsons[1]
        except Exception as e:
            logger.error(f"OpenRouter call or JSON parse failed for model {model}: {e}")
            continue
        all_mode_sections.append(mode_sections_obj)
        all_param_arrays.append(param_array)

    if not all_param_arrays:
        logger.error("No valid parameter suggestions from any model.")
        return None

    merged_param_array = coverage_voting(all_param_arrays)
    final_mode_sections_obj = all_mode_sections[0] if all_mode_sections else {"mode": "", "sections": []}

    suggestion_id = save_optimization_suggestion_to_db(
        db_path, step_id, final_mode_sections_obj, merged_param_array
    )

    # Optionally write to files for backup/debugging
    suggestion_json_save_path = suggestion_json_path if suggestion_json_path else output_path + ".suggestions.json"
    try:
        with open(suggestion_json_save_path, "w", encoding="utf-8") as f:
            json.dump({
                "mode_sections": final_mode_sections_obj,
                "parameters": merged_param_array
            }, f, indent=2)
        logger.info(f"Suggestion JSON written to {suggestion_json_save_path}")
    except Exception as e:
        logger.error(f"Failed to write suggestion JSON: {e}")

    prompt_save_path = suggestion_json_path.replace(".json", ".prompt.txt") if suggestion_json_path else output_path + ".prompt.txt"
    # --- Save prompt to file if path given ---
    try:
        with open(prompt_save_path, "w", encoding="utf-8") as f:
            f.write(prompt)
        logger.info(f"Prompt saved to {prompt_save_path}")
    except Exception as e:
        logger.error(f"Failed to write prompt.txt: {e}")

    try:
        update_parameters(set_path, merged_param_array, output_path)
        logger.info(f"Updated .set file written to: {output_path}")
    except Exception as e:
        logger.error(f"Failed to update .set file: {e}")

    logger.info(f"DB: suggestion_id={suggestion_id}. Updated .set file written to: {output_path}")

    return output_path

def main():
    parser = argparse.ArgumentParser(description="AI-powered .set file optimizer using OpenRouter and multi-model parameter suggestion coverage voting.")
    parser.add_argument("--template", required=True, help="Path to ai_parameter_suggestion_prompt_template file (Markdown)")
    parser.add_argument("--base-parameters", required=True, help="Comma-separated base parameters to highlight for optimization")
    parser.add_argument("--set", required=True, help="Input .set file path")
    parser.add_argument("--spec", required=True, help="PhoenixSpec.csv file path")
    parser.add_argument("--summary", required=True, help="SUMMARY.csv/backtest summary file path")
    parser.add_argument("--openrouter-api-key", required=True, help="OpenRouter API key")
    parser.add_argument("--output", required=True, help="Output .set file path")
    parser.add_argument("--db-path", required=True, help="Path to the SQLite database")
    parser.add_argument("--step-id", required=True, type=int, help="set_file_steps.id to link suggestion to")
    parser.add_argument("--config_xlsx_path", required=False, help="Config.xlsx for performance metrics (optional, for performance criteria)")
    parser.add_argument("--suggestion-json", required=False, help="Output path for suggestion JSON file (default: output.suggestions.json)")
    parser.add_argument("--models", required=False, help="Comma-separated list of OpenRouter model names (e.g. openai/gpt-4o,anthropic/claude-3-opus)")
    args = parser.parse_args()

    models = None
    if args.models:
        models = [m.strip() for m in args.models.split(",") if m.strip()]

    output_path = suggest_mode_and_sections_and_params_openrouter(
        template_path=args.template,
        base_parameters=args.base_parameters,
        set_path=args.set,
        spec_path=args.spec,
        summary_path=args.summary,
        openrouter_api_key=args.openrouter_api_key,
        output_path=args.output,
        db_path=args.db_path,
        step_id=args.step_id,
        config_xlsx_path=args.config_xlsx_path,
        suggestion_json_path=args.suggestion_json,
        models=models,
    )

    if output_path:
        logger.info(f"Optimization completed successfully. Output: {output_path}")
        print(output_path)
    else:
        logger.error("Optimization failed.")

if __name__ == "__main__":
    main()
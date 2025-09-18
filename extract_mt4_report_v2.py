import re
import json
import sys
import logging
#logging.disable(logging.CRITICAL)
logger = logging.getLogger(__name__)

import os
import numpy as np
import sqlite3
import math
import datetime
from bs4 import BeautifulSoup

import build_filename
from build_filename import build_filename
from set_file_updater import update_single_parameter

from wave_analysis import get_wave_analysis_result_block
# --- Logging Setup ---
# class FlushFileHandler(logging.FileHandler):
#     def emit(self, record):
#         super().emit(record)
#         self.flush()
#
# LOG_FILE = os.path.join(os.path.dirname(__file__), "extract_mt4_report.log")
# log_to_file = True  # Set to False to disable file logging
#
# handlers = []
# if log_to_file:
#     handlers.append(FlushFileHandler(LOG_FILE, encoding='utf-8'))
#
# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] %(message)s",
#     handlers=handlers
# )
# logger = logging.getLogger(__name__)

def get_wave_analysis_parameters_from_config(config_xlsx_path, sheet_name="WaveAnalysisConfig"):
    """
    Reads wave analysis parameters from config_xlsx_path (sheet_name), returns a dict for get_wave_analysis_result_block.
    Assumes columns: Name | Value | Description (case-insensitive).
    """
    import openpyxl

    param_names = [
        "csv_path", "depth", "deviation", "backstep", "percentage",
        "force_factor", "normal_wave", "medium_wave", "rare_wave",
        "source"
    ]
    param_dict = {}

    try:
        wb = openpyxl.load_workbook(config_xlsx_path, data_only=True)
        ws = wb[sheet_name]
        # Find Name and Value columns
        header_row = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_idx = header_row.index("name") if "name" in header_row else 0
        value_idx = header_row.index("value") if "value" in header_row else 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[name_idx]:
                continue
            name = str(row[name_idx]).strip()
            value = row[value_idx]
            if name in param_names:
                param_dict[name] = value

        # Convert numeric parameters to correct types
        for k in ["depth", "deviation", "backstep", "percentage", "force_factor", "normal_wave", "medium_wave", "rare_wave"]:
            if k in param_dict and param_dict[k] is not None:
                try:
                    param_dict[k] = float(param_dict[k])
                    if k in ["depth", "backstep"]:
                        param_dict[k] = int(param_dict[k])
                except Exception:
                    pass

        return param_dict

    except Exception as e:
        logger.warning(f"Failed to load wave analysis parameters from config: {e}")
        return {}

def construct_wave_analysis_csv_path(
    data_dir,
    source,
    symbol,
    start_date,
    end_date,
    timeframe
):
    """
    Constructs the full path to the wave analysis CSV file.
    Example:
      data_dir: "C:/Users/Philip/Documents/GitHub/mt4_optimizer/TickData"
      source: "Dukascopy"
      symbol: "AUDCAD"
      start_date: "2006.01.03"
      end_date: "2025.09.05"
      timeframe: "M30"
    Returns:
      "C:/Users/Philip/Documents/GitHub/mt4_optimizer/TickData/Dukascopy-AUDCAD-2006.01.03-2025.09.05-bardata_M30.csv"
    """
    import os

    # Clean up inputs
    data_dir = os.path.normpath(data_dir)
    source = str(source)
    symbol = str(symbol)
    start_date = str(start_date)
    end_date = str(end_date)
    timeframe = str(timeframe).upper()

    filename = f"{source}-{symbol}-{start_date}-{end_date}-bardata_{timeframe}.csv"
    full_path = os.path.join(data_dir, filename)
    return full_path

def read_config_xlsx(path, sheet_name="ai_optimizer"):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        key, value = row
        if key:
            config[str(key)] = str(value) if value is not None else ""
    logger.info(f"Config loaded from {path}: {config}")
    return config

def read_performance_criteria_xlsx(path, sheet_name="performance_criteria"):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    criteria = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        key, value = row[:2]
        if key and value is not None:
            criteria[str(key)] = float(value)
    logger.info(f"Performance criteria loaded from {path}: {criteria}")
    return criteria

# def parse_period_info(period_str):
#     m_period = re.search(r"\(([^)]+)\)", period_str)
#     period = m_period.group(1) if m_period else ""
#     m_start = re.search(r"\)\s+(\d{4}\.\d{2}\.\d{2})", period_str)
#     start_date_str = m_start.group(1) if m_start else ""
#     m_end = re.search(r"\((?:[^\d]+)?(\d{4}\.\d{2}\.\d{2})\s*-\s*(\d{4}\.\d{2}\.\d{2})\)", period_str)
#     end_date_str = m_end.group(2) if m_end else ""
#     start_date = datetime.datetime.strptime(start_date_str, "%Y.%m.%d").date() if start_date_str else None
#     end_date = datetime.datetime.strptime(end_date_str, "%Y.%m.%d").date() if end_date_str else None
#     logger.debug(f"Parsed period info: period={period}, start_date={start_date}, end_date={end_date}")
#     return period, start_date, end_date

def parse_period_info(period_str):
    # PERIOD (e.g., M30)
    m_period = re.search(r"\(([^)]+)\)", period_str)
    period = m_period.group(1) if m_period else ""

    # DATA START DATE (after first parenthesis)
    m_start = re.search(r"\)\s+(\d{4}\.\d{2}\.\d{2})", period_str)
    data_start_date_str = m_start.group(1) if m_start else ""
    data_start_date = datetime.datetime.strptime(data_start_date_str, "%Y.%m.%d").date() if data_start_date_str else None

    # DATA END DATE (before last parenthesis, after dash)
    m_data_end = re.search(r"\)\s+\d{4}\.\d{2}\.\d{2} [\d:]+ - (\d{4}\.\d{2}\.\d{2})", period_str)
    data_end_date_str = m_data_end.group(1) if m_data_end else ""
    data_end_date = datetime.datetime.strptime(data_end_date_str, "%Y.%m.%d").date() if data_end_date_str else None

    # BACKTEST START/END (in last parenthesis)
    m_backtest = re.search(r"\((\d{4}\.\d{2}\.\d{2})\s*-\s*(\d{4}\.\d{2}\.\d{2})\)", period_str)
    backtest_start_date_str = m_backtest.group(1) if m_backtest else ""
    backtest_end_date_str = m_backtest.group(2) if m_backtest else ""
    backtest_start_date = datetime.datetime.strptime(backtest_start_date_str, "%Y.%m.%d").date() if backtest_start_date_str else None
    backtest_end_date = datetime.datetime.strptime(backtest_end_date_str, "%Y.%m.%d").date() if backtest_end_date_str else None

    return period, data_start_date, data_end_date, backtest_start_date, backtest_end_date

def parse_metrics(html_string):
    soup = BeautifulSoup(html_string, "html.parser")
    table = soup.find_all("table")[0]

    def extract_row(key):
        for row in table.find_all("tr"):
            tds = row.find_all("td")
            for idx, td in enumerate(tds):
                val = td.get_text(strip=True)
                if val == key:
                    if idx + 1 < len(tds):
                        return tds[idx + 1].get_text(strip=True)
        return None

    metrics = {}
    metrics["Symbol"] = extract_row("Symbol")
    metrics["Period"] = extract_row("Period")
    metrics["Model"] = extract_row("Model")
    metrics["Parameters"] = extract_row("Parameters")
    metrics["Bars in test"] = extract_row("Bars in test")
    metrics["Ticks modelled"] = extract_row("Ticks modelled")
    metrics["Modelling quality"] = extract_row("Modelling quality")
    metrics["Mismatched charts errors"] = extract_row("Mismatched charts errors")
    metrics["Initial deposit"] = extract_row("Initial deposit")
    metrics["Spread"] = extract_row("Spread")
    metrics["Total net profit"] = extract_row("Total net profit")
    metrics["Gross profit"] = extract_row("Gross profit")
    metrics["Gross loss"] = extract_row("Gross loss")
    metrics["Profit factor"] = extract_row("Profit factor")
    metrics["Expected payoff"] = extract_row("Expected payoff")
    metrics["Absolute drawdown"] = extract_row("Absolute drawdown")
    metrics["Maximal drawdown"] = extract_row("Maximal drawdown")
    metrics["Relative drawdown"] = extract_row("Relative drawdown")
    metrics["Total trades"] = extract_row("Total trades")
    metrics["Short positions (won %)"] = extract_row("Short positions (won %)")
    metrics["Long positions (won %)"] = extract_row("Long positions (won %)")
    metrics["Profit trades (% of total)"] = extract_row("Profit trades (% of total)")
    metrics["Loss trades (% of total)"] = extract_row("Loss trades (% of total)")

    period_str = metrics.get("Period", "")
    period, start_date, end_date, backtest_start_date, backtest_end_date = parse_period_info(period_str)
    metrics["Period"] = period
    metrics["start_date"] = start_date
    metrics["end_date"] = end_date
    metrics["backtest_start_date"] = backtest_start_date
    metrics["backtest_end_date"] = backtest_end_date

    for row in table.find_all("tr"):
        tds = row.find_all("td")
        texts = [td.get_text(strip=True) for td in tds]
        if not texts or len(texts) < 3:
            continue
        first_cell = texts[0].lower()
        if first_cell in ("largest", "average", "maximum", "maximal"):
            label_type = first_cell.capitalize()
            i = 1
            while i + 1 < len(texts):
                label = texts[i]
                value = texts[i + 1]
                key = f"{label_type} {label}"
                metrics[key] = value
                i += 2

    for row in table.find_all("tr"):
        tds = row.find_all("td")
        texts = [td.get_text(strip=True) for td in tds]
        if texts and texts[0].lower() == "average" and "consecutive wins" in texts:
            try:
                win_idx = texts.index("consecutive wins")
                metrics["Max consecutive wins"] = texts[win_idx + 1]
            except Exception:
                pass
            try:
                loss_idx = texts.index("consecutive losses")
                metrics["Max consecutive losses"] = texts[loss_idx + 1]
            except Exception:
                pass

    try:
        profit_trades_str = metrics.get("Profit trades (% of total)", "") or ""
        m = re.match(r"(\d+)\s*\(([\d\.]+)%\)", profit_trades_str)
        metrics["Win rate"] = float(m.group(2)) if m else None
    except Exception:
        metrics["Win rate"] = None

    m = re.match(r"([\d\.\-]+)(?:\s*\(([\d\.]+)%\))?", metrics.get("Maximal drawdown", "") or "")
    metrics["Maximal drawdown value"] = float(m.group(1)) if m else None
    metrics["Maximal drawdown pct"] = float(m.group(2)) if m and m.group(2) else None

    m = re.match(r"([\d\.]+)%\s*\(([\d\.\-]+)\)", metrics.get("Relative drawdown", "") or "")
    metrics["Relative drawdown value"] = float(m.group(2)) if m else None
    metrics["Relative drawdown pct"] = float(m.group(1)) if m else None

    m = re.match(r"(\d+)\s*\(([\d\.]+)%\)", metrics.get("Short positions (won %)", "") or "")
    metrics["Short positions count"] = int(m.group(1)) if m else None
    metrics["Short positions won pct"] = float(m.group(2)) if m else None

    m = re.match(r"(\d+)\s*\(([\d\.]+)%\)", metrics.get("Long positions (won %)", "") or "")
    metrics["Long positions count"] = int(m.group(1)) if m else None
    metrics["Long positions won pct"] = float(m.group(2)) if m else None

    m = re.match(r"(\d+)\s*\(([\d\.\-]+)\)", metrics.get("Maximum consecutive wins (profit in money)", "") or "")
    metrics["Max consecutive wins count"] = int(m.group(1)) if m else None
    metrics["Max consecutive wins profit"] = float(m.group(2)) if m else None

    m = re.match(r"(\d+)\s*\(([\d\.\-]+)\)", metrics.get("Maximum consecutive losses (loss in money)", "") or "")
    metrics["Max consecutive losses count"] = int(m.group(1)) if m else None
    metrics["Max consecutive losses loss"] = float(m.group(2)) if m else None

    m = re.match(r"([\d\.\-]+)\s*\((\d+)\)", metrics.get("Maximal consecutive profit (count of wins)", "") or "")
    metrics["Max consecutive profit"] = float(m.group(1)) if m else None
    metrics["Max consecutive profit count"] = int(m.group(2)) if m else None

    m = re.match(r"([\d\.\-]+)\s*\((\d+)\)", metrics.get("Maximal consecutive loss (count of losses)", "") or "")
    metrics["Max consecutive loss"] = float(m.group(1)) if m else None
    metrics["Max consecutive loss count"] = int(m.group(2)) if m else None

    for k in ["Largest profit trade", "Largest loss trade", "Average profit trade", "Average loss trade"]:
        try:
            metrics[k.lower().replace(" ", "_")] = float(metrics[k])
        except Exception:
            metrics[k.lower().replace(" ", "_")] = None

    logger.info(f"Parsed metrics: {metrics}")
    return metrics

def parse_trades(html_string):
    soup = BeautifulSoup(html_string, "lxml")
    tables = soup.find_all("table")
    if len(tables) < 2:
        logger.warning("Not enough tables found in HTML to parse trades.")
        return []
    trade_table = tables[-1]
    trades = []
    headers = []
    for idx, row in enumerate(trade_table.find_all("tr")):
        tds = row.find_all(["td", "th"])
        row_text = [td.get_text(strip=True) for td in tds]
        if idx == 0:
            headers = row_text
            continue
        if not row_text or not row_text[0].isdigit():
            continue
        trade = {
            "trade_num": int(row_text[0]),
            "time": row_text[1],
            "type": row_text[2],
            "order_id": int(row_text[3]),
            "size": float(row_text[4]),
            "price": float(row_text[5]),
            "sl": float(row_text[6]) if row_text[6] else 0,
            "tp": float(row_text[7]) if row_text[7] else 0,
            "profit": float(row_text[8]) if len(row_text) > 8 and row_text[8] else 0,
            "balance": float(row_text[9]) if len(row_text) > 9 and row_text[9] else 0,
        }
        trades.append(trade)
    logger.info(f"Parsed {len(trades)} trades.")
    return trades

def calculate_sharpe_sortino(trades, initial_deposit):
    returns = []
    prev_balance = initial_deposit
    for trade in trades:
        profit = trade.get("profit", 0)
        if prev_balance > 0 and profit != 0:
            returns.append(profit / prev_balance)
            prev_balance += profit
    if len(returns) < 2:
        return 0.0, 0.0
    mean_return = np.mean(returns)
    std_return = np.std(returns, ddof=1)
    downside_returns = [r for r in returns if r < 0]
    downside_deviation = np.std(downside_returns, ddof=1) if len(downside_returns) > 1 else (abs(downside_returns[0]) if downside_returns else 0.0)
    sharpe_ratio = mean_return / std_return if std_return != 0 else 0.0
    sortino_ratio = mean_return / downside_deviation if downside_deviation != 0 else 0.0
    logger.info(f"Sharpe: {sharpe_ratio}, Sortino: {sortino_ratio}")
    return sharpe_ratio, sortino_ratio

def get_float(val, fallback=0):
    try:
        if isinstance(val, str):
            val = re.sub(r"[^\d\.\-]", "", val)
        return float(val)
    except Exception:
        return fallback

def to_iso(val):
    return val.isoformat() if isinstance(val, (datetime.date, datetime.datetime)) else val

def calculate_custom_metrics(metrics, trades):
    net_profit = get_float(metrics.get("Total net profit"))
    gross_profit = get_float(metrics.get("Gross profit"))
    gross_loss = get_float(metrics.get("Gross loss"))
    profit_factor = get_float(metrics.get("Profit factor"))
    expected_payoff = get_float(metrics.get("Expected payoff"))
    max_drawdown = metrics.get("Maximal drawdown value")
    max_drawdown_pct = metrics.get("Maximal drawdown pct")
    max_relative_drawdown = metrics.get("Relative drawdown value")
    max_relative_drawdown_pct = metrics.get("Relative drawdown pct")
    absolute_drawdown = get_float(metrics.get("Absolute drawdown"))
    initial_deposit = get_float(metrics.get("Initial deposit"))
    total_trades = int(get_float(metrics.get("Total trades")))
    profit_trades_pct = None
    loss_trades_pct = None
    m = re.match(r"(\d+)\s*\(([\d\.]+)%\)", metrics.get("Profit trades (% of total)", ""))
    if m:
        profit_trades_pct = float(m.group(2))
    m = re.match(r"(\d+)\s*\(([\d\.]+)%\)", metrics.get("Loss trades (% of total)", ""))
    if m:
        loss_trades_pct = float(m.group(2))
    largest_profit = get_float(metrics.get("Largest profit trade"))
    largest_loss = get_float(metrics.get("Largest loss trade"))

    sharpe_ratio, sortino_ratio = calculate_sharpe_sortino(trades, initial_deposit)
    recovery_factor = net_profit / max_drawdown if max_drawdown != 0 else 0
    net_profit_per_initial_deposit = net_profit / initial_deposit if initial_deposit != 0 else 0
    absolute_drawdown_per_initial_deposit = absolute_drawdown / initial_deposit if initial_deposit != 0 else 0

    bars_in_test = int(get_float(metrics.get("Bars in test")))
    ticks_modelled = int(get_float(metrics.get("Ticks modelled")))
    modelling_quality = get_float(metrics.get("Modelling quality"))
    mismatched_charts_errors = int(get_float(metrics.get("Mismatched charts errors")))
    spread = get_float(metrics.get("Spread"))
    win_rate = get_float(metrics.get("Win rate"))

    short_positions = metrics.get("Short positions count")
    short_positions_won_pct = metrics.get("Short positions won pct")
    long_positions = metrics.get("Long positions count")
    long_positions_won_pct = metrics.get("Long positions won pct")

    largest_profit_trade = get_float(metrics.get("Largest profit trade"))
    largest_loss_trade = get_float(metrics.get("Largest loss trade"))

    max_consecutive_wins = metrics.get("Max consecutive wins count")
    max_consecutive_wins_profit = metrics.get("Max consecutive wins profit")
    max_consecutive_profit = metrics.get("Max consecutive profit")
    max_consecutive_profit_count = metrics.get("Max consecutive profit count")
    max_consecutive_losses = metrics.get("Max consecutive losses count")
    max_consecutive_losses_loss = metrics.get("Max consecutive losses loss")
    max_consecutive_loss = metrics.get("Max consecutive loss")
    max_consecutive_loss_count = metrics.get("Max consecutive loss count")

    custom = {
        "net_profit": net_profit,
        "gross_profit": gross_profit,
        "gross_loss": gross_loss,
        "profit_factor": profit_factor,
        "expected_payoff": expected_payoff,
        "max_drawdown": max_drawdown,
        "max_drawdown_pct": max_drawdown_pct,
        "max_relative_drawdown": max_relative_drawdown,
        "max_relative_drawdown_pct": max_relative_drawdown_pct,
        "absolute_drawdown": absolute_drawdown,
        "initial_deposit": initial_deposit,
        "total_trades": total_trades,
        "profit_trades_pct": profit_trades_pct,
        "loss_trades_pct": loss_trades_pct,
        "largest_profit": largest_profit,
        "largest_loss": largest_loss,
        "recovery_factor": recovery_factor,
        "sharpe_ratio": 0.0 if math.isnan(sharpe_ratio) else sharpe_ratio,
        "sortino_ratio": 0.0 if math.isnan(sortino_ratio) else sortino_ratio,
        "net_profit_per_initial_deposit": net_profit_per_initial_deposit,
        "absolute_drawdown_per_initial_deposit": absolute_drawdown_per_initial_deposit,
        "bars_in_test": bars_in_test,
        "ticks_modelled": ticks_modelled,
        "modelling_quality": modelling_quality,
        "mismatched_charts_errors": mismatched_charts_errors,
        "spread": spread,
        "win_rate": win_rate,
        "short_positions": short_positions,
        "short_positions_won_pct": short_positions_won_pct,
        "long_positions": long_positions,
        "long_positions_won_pct": long_positions_won_pct,
        "largest_profit_trade": largest_profit_trade,
        "largest_loss_trade": largest_loss_trade,
        "max_consecutive_wins": max_consecutive_wins,
        "max_consecutive_wins_profit": max_consecutive_wins_profit,
        "max_consecutive_profit": max_consecutive_profit,
        "max_consecutive_profit_count": max_consecutive_profit_count,
        "max_consecutive_losses": max_consecutive_losses,
        "max_consecutive_losses_loss": max_consecutive_losses_loss,
        "max_consecutive_loss": max_consecutive_loss,
        "max_consecutive_loss_count": max_consecutive_loss_count,
    }
    logger.info(f"Calculated custom metrics: {custom}")
    return custom

def generate_summary_insert(
    step_id, metric_type, metrics, custom_metrics, parameters_json, summary_csv,
    min_total_recovery, min_trades, min_max_drawdown, criteria_passed, criteria_reason,
    set_file_name, magic_number,
    input_html_file, input_set_file, optimization_pass_id, htm_file_name
):
    sql = """INSERT INTO test_metrics (
    step_id, metric_type, net_profit, gross_profit, gross_loss, profit_factor, expected_payoff,
    max_drawdown, max_drawdown_pct, max_relative_drawdown, max_relative_drawdown_pct,
    absolute_drawdown, initial_deposit, total_trades, profit_trades_pct, loss_trades_pct,
    largest_profit, largest_loss, recovery_factor, sharpe_ratio, sortino_ratio,
    net_profit_per_initial_deposit, absolute_drawdown_per_initial_deposit,
    symbol, period, model, bars_in_test, ticks_modelled, modelling_quality, mismatched_charts_errors,
    spread, short_positions, short_positions_won_pct, long_positions, long_positions_won_pct,
    largest_profit_trade, largest_loss_trade, max_consecutive_wins, max_consecutive_wins_profit,
    max_consecutive_profit, max_consecutive_profit_count, max_consecutive_losses, max_consecutive_losses_loss,
    max_consecutive_loss, max_consecutive_loss_count, win_rate,
    metrics_json, parameters_json, summary_csv,
    start_date, end_date, min_total_recovery, min_trades, min_max_drawdown,
    criteria_passed, criteria_reason, set_file_name, magic_number,
    input_html_file, input_set_file, optimization_pass_id, htm_file_name
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""
    values = [
        step_id,
        metric_type,
        custom_metrics["net_profit"],
        custom_metrics["gross_profit"],
        custom_metrics["gross_loss"],
        custom_metrics["profit_factor"],
        custom_metrics["expected_payoff"],
        custom_metrics["max_drawdown"],
        custom_metrics["max_drawdown_pct"],
        custom_metrics["max_relative_drawdown"],
        custom_metrics["max_relative_drawdown_pct"],
        custom_metrics["absolute_drawdown"],
        custom_metrics["initial_deposit"],
        custom_metrics["total_trades"],
        custom_metrics["profit_trades_pct"],
        custom_metrics["loss_trades_pct"],
        custom_metrics["largest_profit"],
        custom_metrics["largest_loss"],
        custom_metrics["recovery_factor"],
        custom_metrics["sharpe_ratio"],
        custom_metrics["sortino_ratio"],
        custom_metrics["net_profit_per_initial_deposit"],
        custom_metrics["absolute_drawdown_per_initial_deposit"],
        metrics.get("Symbol"),
        metrics.get("Period"),
        metrics.get("Model"),
        custom_metrics["bars_in_test"],
        custom_metrics["ticks_modelled"],
        custom_metrics["modelling_quality"],
        custom_metrics["mismatched_charts_errors"],
        custom_metrics["spread"],
        custom_metrics["short_positions"],
        custom_metrics["short_positions_won_pct"],
        custom_metrics["long_positions"],
        custom_metrics["long_positions_won_pct"],
        custom_metrics["largest_profit_trade"],
        custom_metrics["largest_loss_trade"],
        custom_metrics["max_consecutive_wins"],
        custom_metrics["max_consecutive_wins_profit"],
        custom_metrics["max_consecutive_profit"],
        custom_metrics["max_consecutive_profit_count"],
        custom_metrics["max_consecutive_losses"],
        custom_metrics["max_consecutive_losses_loss"],
        custom_metrics["max_consecutive_loss"],
        custom_metrics["max_consecutive_loss_count"],
        custom_metrics["win_rate"],
        json.dumps(custom_metrics),
        parameters_json,
        summary_csv,
        to_iso(metrics.get("start_date")),
        to_iso(metrics.get("end_date")),
        min_total_recovery,
        min_trades,
        min_max_drawdown,
        int(bool(criteria_passed)),
        criteria_reason,
        set_file_name,
        magic_number,
        input_html_file,
        input_set_file,
        optimization_pass_id,
        htm_file_name
    ]
    logger.debug(f"SQL Insert: {sql}")
    logger.debug(f"SQL Values: {values}")
    return sql, values

def generate_trade_inserts(step_id, trades, symbol, magic_number=0):
    sql = """
INSERT INTO trades (
    step_id, trade_num, time, type, order_id, size, price, sl, tp, profit, balance, drawdown, comment, symbol, magic_number, ticket
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""
    inserts = []
    for trade in trades:
        inserts.append([
            step_id,
            trade["trade_num"],
            trade["time"],
            trade["type"],
            trade["order_id"],
            trade["size"],
            trade["price"],
            trade["sl"],
            trade["tp"],
            trade.get("profit", 0),
            trade.get("balance", 0),
            None,
            None,
            symbol,
            magic_number,
            None,
        ])
    logger.debug(f"Generated {len(inserts)} trade inserts.")
    return sql, inserts

def gen_summary_csv(metrics, custom_metrics, csv_name):
    ordered_keys = [k for k in custom_metrics if k != "Parameters"] + \
                   [k for k in metrics if k not in custom_metrics and k != "Parameters"]

    header_row = ",".join(ordered_keys)
    value_row = ",".join([str(custom_metrics.get(key, metrics.get(key, ""))) for key in ordered_keys])

    summary_csv = header_row + "\n" + value_row

    with open(csv_name, 'w', encoding='utf-8') as f:
        f.write(summary_csv)
    logger.info(f"Summary CSV written to {csv_name}")
    return summary_csv

def calc_drawdown_sl_money(max_drawdown, step=100):
    quotient = int(max_drawdown // step)
    remainder = max_drawdown % step
    base = (quotient + 1) * step  # default to next step
    if remainder > step / 2:
        base += step  # bump up one more step
    return base

def copy_and_rename_html_and_gif(input_html, output_set_file_name, output_dir):
    import shutil
    from pathlib import Path
    import re

    # Determine output HTML and GIF names
    base_name = Path(output_set_file_name).stem  # e.g. ...S358
    out_html_name = base_name + ".htm"
    out_gif_name = base_name + ".gif"
    out_html_path = Path(output_dir) / out_html_name
    out_gif_path = Path(output_dir) / out_gif_name

    # Read HTML, find src GIF
    with open(input_html, "r", encoding="utf-8") as f:
        html = f.read()

    # Find original GIF name(s)
    gif_matches = re.findall(r'<img\s+src="([^"]+\.gif)"', html, re.IGNORECASE)
    if not gif_matches:
        raise Exception("No GIF image found in HTML!")
    orig_gif_name = gif_matches[0]
    orig_gif_path = Path(input_html).parent / orig_gif_name

    # Replace all GIF src with new GIF name
    html_new = re.sub(r'(<img\s+src=")[^"]+\.gif(")', r'\1' + out_gif_name + r'\2', html, flags=re.IGNORECASE)

    # Write new HTML
    with open(out_html_path, "w", encoding="utf-8") as f:
        f.write(html_new)

    # Copy GIF
    shutil.copy(orig_gif_path, out_gif_path)

    # Return both names and full paths for downstream use
    return (
        out_html_name,                # e.g. MyTestS358.htm
        str(out_html_path),           # full path to new HTML
        out_gif_name,                 # e.g. MyTestS358.gif
        str(out_gif_path),            # full path to new GIF
        str(orig_gif_path)            # full path to original GIF (input artifact)
    )

def insert_set_file_artifacts(
    db_path,
    step_id,
    artifact_type=None,
    file_path=None,
    meta_json=None,
    link_type=None,
    link_id=None
):
    """
    Insert a row into set_file_artifacts table.
    Reads file_blob from file_path if provided and file exists.
    Arguments:
        db_path (str): Path to the SQLite database file.
        step_id (int): Foreign key referencing set_file_steps(id).
        artifact_type (str or None)
        file_path (str or None)
        meta_json (str or None)
        link_type (str or None)
        link_id (int or None)
    Returns:
        int: The id of the newly inserted row.
    """
    # Read file_blob from file_path, if it exists
    file_blob = None
    if file_path and os.path.isfile(file_path):
        with open(file_path, "rb") as f:
            file_blob = f.read()

    sql = """
        INSERT INTO set_file_artifacts (
            step_id, artifact_type, file_path, meta_json, file_blob, link_type, link_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """
    params = (step_id, artifact_type, file_path, meta_json, file_blob, link_type, link_id)
    conn = sqlite3.connect(db_path)
    #conn.execute("PRAGMA key = 'Kh78784bt!'")
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()

def process_mt4_report(
    html_file,
    step_id,
    metric_type,
    EA_name,
    input_set_file,
    output_set_file_path=r"C:\Users\Philip\Documents\GitHub\EA_Automation\02_backtest\\",
    db_path="C:\\Users\\Philip\\Documents\\GitHub\\EA_Automation\\EA_Automation.db",
    summary_metrics_path="summary_metrics.csv",
    config_xlsx_path=None,
    perf_criteria_xlsx_path=None,
    optimization_pass_id=None
):
    logger.info(f"Processing MT4 report: {html_file}, step_id={step_id}, metric_type={metric_type}, EA_name={EA_name}")
    config = read_config_xlsx(config_xlsx_path) if config_xlsx_path else None
    perf_criteria = read_performance_criteria_xlsx(perf_criteria_xlsx_path) if perf_criteria_xlsx_path else None

    import shutil
    with open(html_file, encoding="utf-8") as f:
        html = f.read()

    metrics = parse_metrics(html)
    trades = parse_trades(html)
    custom_metrics = calculate_custom_metrics(metrics, trades)

    max_drawdown_val = custom_metrics["max_drawdown"]
    drawdown_sl_money = None
    if max_drawdown_val is not None and max_drawdown_val > 0:
        drawdown_sl_money = calc_drawdown_sl_money(float(max_drawdown_val), step=100)
    logger.info(f"DrawDown_SL_Money calculated: {drawdown_sl_money}")

    if perf_criteria is not None:
        min_total_recovery = perf_criteria.get("min_total_recovery", 0)
        min_trades = perf_criteria.get("min_trades", 0)
        min_max_drawdown = perf_criteria.get("min_max_drawdown", float('inf'))
    else:
        min_total_recovery = 0
        min_trades = 0
        min_max_drawdown = float('inf')

    recovery_factor = custom_metrics["recovery_factor"]
    total_trades = custom_metrics["total_trades"]
    max_drawdown_val = custom_metrics["max_drawdown"]

    criteria_reason_list = []
    criteria_passed = True
    if recovery_factor < min_total_recovery:
        criteria_reason_list.append("recovery_factor")
        criteria_passed = False
    if total_trades < min_trades:
        criteria_reason_list.append("total_trades")
        criteria_passed = False
    if max_drawdown_val > min_max_drawdown:
        criteria_reason_list.append("max_drawdown")
        criteria_passed = False
    criteria_reason = "All passed" if not criteria_reason_list else ", ".join(criteria_reason_list)
    logger.info(f"Performance criteria: passed={criteria_passed}, reason={criteria_reason}")

    parameters_json = json.dumps({"Parameters": metrics.get("Parameters", "")})
    metrics_no_parameters = {k: v for k, v in metrics.items() if k != "Parameters"}

    Symbol = metrics.get("Symbol", "")
    Timeframe = metrics.get("period", "") or metrics.get("Period", "")
    InitialDeposit = metrics.get("Initial deposit", "")
    ProfitAmount = metrics.get("Total net profit", "")
    DrawDown = metrics.get("Maximal drawdown value", "")
    StartDate = metrics.get("start_date")
    EndDate = metrics.get("end_date")
    if StartDate:
        StartDate = StartDate.strftime("%Y%m%d")
    else:
        StartDate = ""
    if EndDate:
        EndDate = EndDate.strftime("%Y%m%d")
    else:
        EndDate = ""
    Stoploss = str(drawdown_sl_money) if drawdown_sl_money is not None else ""
    WinRate = str(metrics.get("Win rate", ""))
    ProfitFactor = metrics.get("Profit factor", "")
    NumTrade = metrics.get("Total trades", "")
    SetVersion = "1"
    Step = step_id

    output_set_file_name, magic_number = build_filename(
        EA=EA_name,
        Symbol=Symbol,
        Timeframe=Timeframe,
        InitialDeposit=InitialDeposit,
        ProfitAmount=ProfitAmount,
        DrawDown=DrawDown,
        StartDate=StartDate,
        EndDate=EndDate,
        Stoploss=Stoploss,
        WinRate=WinRate,
        ProfitFactor=ProfitFactor,
        NumTrade=NumTrade,
        SetVersion=SetVersion,
        Step=Step
    )
    output_file = os.path.join(output_set_file_path, output_set_file_name)

    output_set_name_no_ext = os.path.splitext(output_set_file_name)[0]
    summary_metrics_full_path = os.path.join(
        output_set_file_path,
        f"{output_set_name_no_ext}_{summary_metrics_path.lstrip('_')}"
    )

    summary_csv = gen_summary_csv(metrics_no_parameters, custom_metrics, summary_metrics_full_path)

    shutil.copy(input_set_file, output_file)
    logger.info(f"Copied {input_set_file} to {output_file}")
    if drawdown_sl_money is not None:
        update_single_parameter(output_file, "DrawDown_SL_Money", drawdown_sl_money)
        logger.info(f"Updated DrawDown_SL_Money in {output_file} to {drawdown_sl_money}")
    if magic_number is not None:
        update_single_parameter(output_file, "Magic", magic_number)
        logger.info(f"Updated Magic in {output_file} to {magic_number}")

    out_html_name, out_html_path, out_gif_name, out_gif_path, orig_gif_path = copy_and_rename_html_and_gif(
        html_file, output_set_file_name, output_set_file_path
    )

    summary_sql, summary_values = generate_summary_insert(
        step_id, metric_type, metrics, custom_metrics, parameters_json, summary_csv,
        min_total_recovery, min_trades, min_max_drawdown, criteria_passed, criteria_reason,
        output_set_file_name, magic_number,
        html_file, input_set_file, optimization_pass_id, out_html_name
    )
    trade_sql, trade_values = generate_trade_inserts(
        step_id, trades, metrics.get("Symbol", "")
    )

    connection = sqlite3.connect(db_path)
    #connection.execute("PRAGMA key = 'Kh78784bt!'")
    cursor = connection.cursor()
    try:
        cursor.execute(summary_sql, summary_values)
        # Get the test_metrics id (lastrowid from the test_metrics insert)
        test_metrics_id = cursor.lastrowid
        logger.info("Inserted test_metrics summary.")
        for trade in trade_values:
            cursor.execute(trade_sql, trade)
        logger.info(f"Inserted {len(trade_values)} trades.")
        connection.commit()
        logger.info("Database commit successful.")
    except Exception as e:
        logger.exception(f"Database operation failed: {e}")
        raise
    finally:
        cursor.close()
        connection.close()
        logger.info("Database connection closed.")

    # Prepare artifact files to record
    artifact_files = [
        # Input files
        {"artifact_type": "input_html", "file_path": html_file},
        {"artifact_type": "input_set", "file_path": input_set_file},
        # GIF in HTML (input artifact)
        {"artifact_type": "input_gif", "file_path": orig_gif_path},
        # Output/generated files
        {"artifact_type": "output_set", "file_path": output_file},
        {"artifact_type": "summary_metrics_csv", "file_path": summary_metrics_full_path},
        {"artifact_type": "output_html", "file_path": out_html_path},
        {"artifact_type": "output_gif", "file_path": out_gif_path},
    ]

    if config is not None:
        try:
            from ai_set_optimizer_openrouter import suggest_mode_and_sections_and_params_openrouter

            template_path = config.get("template_path")
            base_parameters = config.get("base_parameters", "")
            spec_path = config.get("spec_path")
            api_key = config.get("api_key")
            models = config.get("models")

            set_path = output_file
            summary_path = summary_metrics_full_path
            output_path = output_file.replace(".set", "-AI-Suggest-Opt.set")
            suggestion_json_path = output_file.replace(".set", "-AI-Suggestion.json")
            db_path = db_path
            step_id = step_id

            # (A) Get wave analysis parameters from config
            wave_params = get_wave_analysis_parameters_from_config(config_xlsx_path, sheet_name="WaveAnalysisConfig")

            Symbol = metrics.get("Symbol", "")
            Symbol_code = Symbol.split()[0]  # Gets "AUDCAD"
            Timeframe = metrics.get("period", "") or metrics.get("Period", "")
            StartDate = metrics.get("backtest_start_date")
            EndDate = metrics.get("backtest_end_date")

            # Format start/end dates for filename convention
            if StartDate:
                start_date_str = StartDate.strftime("%Y.%m.%d")
            else:
                start_date_str = ""
            if EndDate:
                end_date_str = EndDate.strftime("%Y.%m.%d")
            else:
                end_date_str = ""

            # (B) Construct the CSV path
            wave_params["csv_path"] = construct_wave_analysis_csv_path(
                data_dir=wave_params["csv_path"],  # (or wave_params["data_dir"] if that's your folder)
                source=wave_params["source"],
                symbol=Symbol_code,
                start_date=start_date_str,
                end_date=end_date_str,
                timeframe=Timeframe
            )

            # (C) Build the wave analysis block
            wave_analysis_block = get_wave_analysis_result_block(
                csv_path=wave_params["csv_path"],
                depth=wave_params["depth"],
                deviation=wave_params["deviation"],
                backstep=wave_params["backstep"],
                percentage=wave_params["percentage"],
                force_factor=wave_params["force_factor"],
                normal_wave=wave_params["normal_wave"],
                medium_wave=wave_params["medium_wave"]
            )

            ai_set_file_path = suggest_mode_and_sections_and_params_openrouter(
                template_path=template_path,
                base_parameters=base_parameters,
                set_path=set_path,
                spec_path=spec_path,
                summary_path=summary_path,
                openrouter_api_key=api_key,
                output_path=output_path,
                db_path=db_path,
                step_id=step_id,
                config_xlsx_path=config_xlsx_path,
                suggestion_json_path=suggestion_json_path,
                models=models,
                wave_analysis_block=wave_analysis_block
            )
            logger.info(f"AI set file suggestion complete: {output_path}")

            # Optionally include AI files, only if config is not None (AI suggestion runs)
            if config is not None:
                artifact_files.extend([
                    {"artifact_type": "ai_set", "file_path": output_path},
                    {"artifact_type": "ai_json", "file_path": suggestion_json_path},
                    {"artifact_type": "ai_prompt", "file_path": suggestion_json_path.replace(".json", ".prompt.txt")},
                ])

            for artifact in artifact_files:
                if artifact["file_path"] and os.path.exists(artifact["file_path"]):
                    insert_set_file_artifacts(
                        db_path=db_path,
                        step_id=step_id,
                        artifact_type=artifact["artifact_type"],
                        file_path=artifact["file_path"],
                        meta_json=json.dumps({}),  # Add metadata if needed
                        link_type="test_metrics",
                        link_id=test_metrics_id,
                    )
            return json.dumps({"ai_set_file_path": output_path})
        except Exception as e:
            logger.exception(f"AI set file suggestion failed: {e}")
            return json.dumps({"result": "ai_set_file_suggestion_failed"})
    else:
        for artifact in artifact_files:
            if artifact["file_path"] and os.path.exists(artifact["file_path"]):
                insert_set_file_artifacts(
                    db_path=db_path,
                    step_id=step_id,
                    artifact_type=artifact["artifact_type"],
                    file_path=artifact["file_path"],
                    meta_json=json.dumps({}),  # Add metadata if needed
                    link_type="test_metrics",
                    link_id=test_metrics_id,
                )
        result = "success"
        logger.info("Processing complete.")
        return json.dumps({"result": result})

# if __name__ == "__main__":
#     import argparse
#     parser = argparse.ArgumentParser(description="Process MT4 Backtest Report HTML and insert results into DB.")
#     parser.add_argument("html_file", help="Path to the MT4 backtest report HTML file")
#     parser.add_argument("--step_id", type=int, required=True, help="Step ID for the report")
#     parser.add_argument("--metric_type", type=str, required=True, help="Type of metrics/report")
#     parser.add_argument("--EA_name", type=str, required=True, help="EA name for filename")
#     parser.add_argument("--input_set_file", type=str, required=True, help="Input .set file to copy and update")
#     parser.add_argument("--output_set_file_path", type=str, default=r"C:\Users\Philip\Documents\GitHub\EA_Automation\02_backtest\\", help="Directory to save output set file")
#     parser.add_argument("--db_path", type=str, default=r"C:\Users\Philip\Documents\GitHub\EA_Automation\EA_Automation.db", help="Path to SQLite database")
#     parser.add_argument("--summary_metrics_path", type=str, default="_summary_metrics.csv", help="Path to output summary metrics CSV file")
#     parser.add_argument("--config_xlsx", type=str, required=False, help="Path to config.xlsx")
#     parser.add_argument("--perf_criteria_xlsx", type=str, required=False, help="Path to performance_criteria.xlsx")
#     parser.add_argument("--optimization_pass_id", type=int, required=False, help="Optimization pass ID")
#     args = parser.parse_args()
#
#     process_mt4_report(
#         args.html_file,
#         args.step_id,
#         args.metric_type,
#         args.EA_name,
#         args.input_set_file,
#         output_set_file_path=args.output_set_file_path,
#         db_path=args.db_path,
#         summary_metrics_path=args.summary_metrics_path,
#         config_xlsx_path=args.config_xlsx,  # Pass the path, not dict
#         perf_criteria_xlsx_path=args.perf_criteria_xlsx,
#         optimization_pass_id = args.optimization_pass_id  # Optional, can be None
#     )

# New Main to output JSON with success/error/data and prepare for pyinstall packaging for integration with uipath
# if __name__ == "__main__":
#     import argparse
#     import json
#     output = {}
#     try:
#         parser = argparse.ArgumentParser(description="Process MT4 Backtest Report HTML and insert results into DB.")
#         parser.add_argument("html_file", help="Path to the MT4 backtest report HTML file")
#         parser.add_argument("--step_id", type=int, required=True, help="Step ID for the report")
#         parser.add_argument("--metric_type", type=str, required=True, help="Type of metrics/report")
#         parser.add_argument("--EA_name", type=str, required=True, help="EA name for filename")
#         parser.add_argument("--input_set_file", type=str, required=True, help="Input .set file to copy and update")
#         parser.add_argument("--output_set_file_path", type=str, default=r"C:\Users\Philip\Documents\GitHub\EA_Automation\02_backtest\\", help="Directory to save output set file")
#         parser.add_argument("--db_path", type=str, default=r"C:\Users\Philip\Documents\GitHub\EA_Automation\EA_Automation.db", help="Path to SQLite database")
#         parser.add_argument("--summary_metrics_path", type=str, default="_summary_metrics.csv", help="Path to output summary metrics CSV file")
#         parser.add_argument("--config_xlsx", type=str, required=False, help="Path to config.xlsx")
#         parser.add_argument("--perf_criteria_xlsx", type=str, required=False, help="Path to performance_criteria.xlsx")
#         parser.add_argument("--optimization_pass_id", type=int, required=False, help="Optimization pass ID")
#         args = parser.parse_args()
#
#         result = process_mt4_report(
#             args.html_file,
#             args.step_id,
#             args.metric_type,
#             args.EA_name,
#             args.input_set_file,
#             output_set_file_path=args.output_set_file_path,
#             db_path=args.db_path,
#             summary_metrics_path=args.summary_metrics_path,
#             config_xlsx_path=args.config_xlsx,
#             perf_criteria_xlsx_path=args.perf_criteria_xlsx,
#             optimization_pass_id=args.optimization_pass_id
#         )
#         output["success"] = True
#         output["error"] = ""
#         try:
#             result_dict = json.loads(result)
#             if isinstance(result_dict, dict):
#                 output.update(result_dict)  # Flatten keys into output
#             else:
#                 output["result"] = result_dict
#         except Exception:
#             output["result"] = result
#     except Exception as e:
#         output["success"] = False
#         output["error"] = str(e)
#     print(json.dumps(output))

#remove the argparse parser and use direct sys.argv index-based argument parsing, so your script will accept arguments in a strict positional order, making it compatible with PowerShell's --% operator (which simply passes all arguments as-is to the EXE).
if __name__ == "__main__":
    import sys
    import json

    output = {}
    try:
        # Remove '--%' if present from PowerShell
        if '--%' in sys.argv:
            sys.argv.remove('--%')

        # Usage: html_file step_id metric_type EA_name input_set_file [output_set_file_path] [db_path] [summary_metrics_path] [config_xlsx] [perf_criteria_xlsx] [optimization_pass_id]
        if len(sys.argv) < 6:
            print(json.dumps({
                "success": False,
                "error": "Insufficient arguments. Usage: html_file step_id metric_type EA_name input_set_file [output_set_file_path] [db_path] [summary_metrics_path] [config_xlsx] [perf_criteria_xlsx] [optimization_pass_id]"
            }))
            sys.exit(1)

        html_file = sys.argv[1]
        step_id = int(sys.argv[2])
        metric_type = sys.argv[3]
        EA_name = sys.argv[4]
        input_set_file = sys.argv[5]
        output_set_file_path = sys.argv[6] if len(sys.argv) > 6 else r"C:\Users\Philip\Documents\GitHub\EA_Automation\02_backtest\\"
        db_path = sys.argv[7] if len(sys.argv) > 7 else r"C:\Users\Philip\Documents\GitHub\EA_Automation\EA_Automation.db"
        summary_metrics_path = sys.argv[8] if len(sys.argv) > 8 else "_summary_metrics.csv"
        config_xlsx = sys.argv[9] if len(sys.argv) > 9 else None
        perf_criteria_xlsx = sys.argv[10] if len(sys.argv) > 10 else None
        optimization_pass_id = int(sys.argv[11]) if len(sys.argv) > 11 else None

        result = process_mt4_report(
            html_file,
            step_id,
            metric_type,
            EA_name,
            input_set_file,
            output_set_file_path=output_set_file_path,
            db_path=db_path,
            summary_metrics_path=summary_metrics_path,
            config_xlsx_path=config_xlsx,
            perf_criteria_xlsx_path=perf_criteria_xlsx,
            optimization_pass_id=optimization_pass_id
        )
        output["success"] = True
        output["error"] = ""
        try:
            result_dict = json.loads(result)
            if isinstance(result_dict, dict):
                output.update(result_dict)
            else:
                output["result"] = result_dict
        except Exception:
            output["result"] = result
    except Exception as e:
        output["success"] = False
        output["error"] = str(e)
    print(json.dumps(output))
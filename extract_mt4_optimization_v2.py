import sqlite3
import re
import json
import os
import argparse
from bs4 import BeautifulSoup
import openpyxl  # For reading config.xlsx
from pathlib import Path

def get_ea_name_from_title(soup):
    title_tag = soup.find('title')
    if title_tag and 'Strategy Tester:' in title_tag.text:
        return title_tag.text.split('Strategy Tester:')[-1].strip()
    return ""

def get_metadata_value(rows, label):
    for row in rows:
        tds = row.find_all('td')
        for td in tds:
            if label in td.text:
                return tds[-1].text.strip() if len(tds) > 1 else ""
    return ""

def safe_float(val):
    try:
        return float(val.replace(',', '').replace('\xa0', '').strip())
    except Exception:
        return 0.0

def safe_int(val):
    try:
        return int(val.replace(',', '').replace('\xa0', '').strip())
    except Exception:
        return 0

def read_performance_criteria_xlsx(path, sheet_name="performance_criteria"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    criteria = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        key, value = row[:2]
        if key and value is not None:
            criteria[str(key)] = float(value)
    return criteria

def read_optimization_config(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_weights = wb['optimization_weights']
    ws_setting = wb['optimization_setting']
    weights = {}
    for row in ws_weights.iter_rows(min_row=2, values_only=True):
        metric, weight = row[:2]
        if metric and weight is not None:
            weights[str(metric)] = float(weight)
    # Set defaults
    top_n = 10
    fuzzy = 0.9
    distance = 0.1
    # Read settings from optimization_setting
    for row in ws_setting.iter_rows(min_row=2, values_only=True):
        key, value = row[:2]
        if key == 'top_n' and value is not None:
            top_n = int(value)
        elif key == 'fuzzy' and value is not None:
            fuzzy = float(value)
        elif key == 'distance' and value is not None:
            distance = float(value)
    return weights, top_n, fuzzy, distance

def parse_report(html_path):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'lxml')  # just this line changed!

    ea_name = get_ea_name_from_title(soup)
    if not ea_name:
        ea_div = soup.find('div', string=re.compile("Ace Phoenix|Ace Falcon"))
        ea_name = ea_div.text.strip() if ea_div else ""

    account_div = soup.find('div', string=re.compile("VantageInternational"))
    mt4_account = account_div.text.strip() if account_div else ""

    meta_table = soup.find_all('table')[0]
    rows = meta_table.find_all('tr')

    symbol = get_metadata_value(rows, "Symbol")
    period = get_metadata_value(rows, "Period")
    model = get_metadata_value(rows, "Model")
    initial_deposit = safe_float(get_metadata_value(rows, "Initial deposit"))
    spread = safe_float(get_metadata_value(rows, "Spread"))

    date_range = ""
    if period and '(' in period and ')' in period:
        matches = re.findall(r'\(([^()]*)\)', period)
        if matches:
            date_range = matches[-1]
            period = period.split('(')[0].strip()

    passes_table = None
    tables = soup.find_all('table')
    if len(tables) > 1:
        passes_table = tables[1]
    passes_rows = passes_table.find_all('tr', attrs={'align': 'right'}) if passes_table else []

    passes = []
    for tr in passes_rows:
        tds = tr.find_all('td')
        if len(tds) < 7:
            continue  # skip malformed rows
        try:
            pass_number = safe_int(tds[0].text)
            profit = safe_float(tds[1].text)
            total_trades = safe_int(tds[2].text)
            profit_factor = safe_float(tds[3].text)
            expected_payoff = safe_float(tds[4].text)
            drawdown_abs = safe_float(tds[5].text)
            drawdown_pct = safe_float(tds[6].text)
        except Exception:
            continue
        parameters_json = tds[0]['title'] if tds[0].has_attr('title') else ""
        passes.append({
            'pass_number': pass_number,
            'profit': profit,
            'total_trades': total_trades,
            'profit_factor': profit_factor,
            'expected_payoff': expected_payoff,
            'drawdown_abs': drawdown_abs,
            'drawdown_pct': drawdown_pct,
            'parameters_json': parameters_json,
        })

    report = {
        'ea_name': ea_name,
        'mt4_account': mt4_account,
        'symbol': symbol,
        'period': period,
        'date_range': date_range,
        'model': model,
        'initial_deposit': initial_deposit,
        'spread': spread,
        'passes': passes
    }
    return report

def evaluate_pass(p, criteria):
    # Use defaults if criteria is missing
    min_total_recovery = criteria.get('min_total_recovery', 3)
    min_trades = int(criteria.get('min_trades', 300))
    min_max_drawdown = criteria.get('min_max_drawdown', 1200)
    recovery_factor = p['profit'] / p['drawdown_abs'] if p['drawdown_abs'] else 0.0
    score = recovery_factor * p['profit_factor'] if p['drawdown_abs'] else 0.0

    criteria_passed = True
    reasons = []
    if recovery_factor < min_total_recovery:
        criteria_passed = False
        reasons.append("recovery_factor")
    if p['total_trades'] < min_trades:
        criteria_passed = False
        reasons.append("total_trades")
    if p['drawdown_abs'] > min_max_drawdown:
        criteria_passed = False
        reasons.append("max_drawdown")
    criteria_reason = "All passed" if not reasons else ", ".join(reasons)

    return {
        'recovery_factor': recovery_factor,
        'score': score,
        'criteria_passed': int(criteria_passed),
        'criteria_reason': criteria_reason,
        'min_total_recovery': min_total_recovery,
        'min_trades': min_trades,
        'min_max_drawdown': min_max_drawdown,
    }

def insert_into_db(report, db_path, step_id, criteria):
    conn = sqlite3.connect(db_path)
    #conn.execute("PRAGMA key = 'Kh78784bt!'")
    cur = conn.cursor()
    # Insert report
    cur.execute("""
        INSERT INTO optimization_reports 
        (step_id, ea_name, mt4_account, symbol, period, date_range, model, initial_deposit, spread, passes_count)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (step_id, report['ea_name'], report['mt4_account'], report['symbol'], report['period'], report['date_range'],
          report['model'], report['initial_deposit'], report['spread'], len(report['passes'])))
    report_id = cur.lastrowid

    # Insert all passes (with evaluation)
    for p in report['passes']:
        evals = evaluate_pass(p, criteria)
        pass_metrics = {
            "recovery_factor": evals["recovery_factor"],
            "score": evals["score"],
            "profit_factor": p["profit_factor"],
            "drawdown_abs": p["drawdown_abs"]
            # Add more metrics if needed
        }
        cur.execute("""
            INSERT INTO optimization_passes
            (report_id, pass_number, profit, total_trades, profit_factor, expected_payoff, drawdown_abs, drawdown_pct, parameters_json, pass_metrics_json, score,
             min_total_recovery, min_trades, min_max_drawdown, criteria_passed, criteria_reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (report_id, p['pass_number'], p['profit'], p['total_trades'], p['profit_factor'], p['expected_payoff'],
              p['drawdown_abs'], p['drawdown_pct'], p['parameters_json'], json.dumps(pass_metrics),
              evals['score'], evals['min_total_recovery'], evals['min_trades'], evals['min_max_drawdown'],
              evals['criteria_passed'], evals['criteria_reason']))

    conn.commit()
    conn.close()
    return report_id

# def get_top_n_passes(db_path, report_id, weights, top_n, fuzzy_threshold=0.9):
#     """
#     Retrieve top N optimization passes based on weighted_score, including all passes
#     within fuzzy_threshold * max(weighted_score). Guarantees at least one pass if any exist.
#     Ensures passes are distinct among key performance metrics.
#     """
#     import sqlite3
#     w_profit = weights.get('profit', 0.3)
#     w_total_recovery = weights.get('total_recovery', 0.3)
#     w_drawdown_abs = weights.get('drawdown_abs', -0.2)
#     w_total_trades = weights.get('total_trades', 0.2)
#     sql = f"""
#     WITH stats AS (
#         SELECT
#             report_id,
#             MIN(profit) AS min_profit,
#             MAX(profit) AS max_profit,
#             MIN(drawdown_abs) AS min_drawdown,
#             MAX(drawdown_abs) AS max_drawdown,
#             MIN(profit/drawdown_abs) AS min_total_recovery,
#             MAX(profit/drawdown_abs) AS max_total_recovery,
#             MIN(total_trades) AS min_trades,
#             MAX(total_trades) AS max_trades
#         FROM optimization_passes
#         WHERE report_id = ?
#         GROUP BY report_id
#     ),
#     scored AS (
#         SELECT
#             p.*,
#             (p.profit - s.min_profit) / NULLIF(s.max_profit - s.min_profit, 0) AS norm_profit,
#             ((p.profit / NULLIF(p.drawdown_abs,0)) - s.min_total_recovery) / NULLIF(s.max_total_recovery - s.min_total_recovery, 0) AS norm_total_recovery,
#             (p.drawdown_abs - s.min_drawdown) / NULLIF(s.max_drawdown - s.min_drawdown, 0) AS norm_drawdown,
#             (p.total_trades - s.min_trades) / NULLIF(s.max_trades - s.min_trades, 0) AS norm_trades,
#             (
#                 {w_profit} * (p.profit - s.min_profit) / NULLIF(s.max_profit - s.min_profit, 0)
#                 + {w_total_recovery} * ((p.profit / NULLIF(p.drawdown_abs,0)) - s.min_total_recovery) / NULLIF(s.max_total_recovery - s.min_total_recovery, 0)
#                 + {w_drawdown_abs} * (p.drawdown_abs - s.min_drawdown) / NULLIF(s.max_drawdown - s.min_drawdown, 0)
#                 + {w_total_trades} * (p.total_trades - s.min_trades) / NULLIF(s.max_trades - s.min_trades, 0)
#             ) AS weighted_score
#         FROM optimization_passes p
#         JOIN stats s ON s.report_id = p.report_id
#         WHERE p.report_id = ?
#     ),
#     deduped AS (
#         SELECT
#             MIN(id) AS id
#         FROM scored
#         GROUP BY profit, drawdown_abs, total_trades, weighted_score
#     )
#     SELECT s.*
#     FROM scored s
#     JOIN deduped d ON s.id = d.id
#     ORDER BY s.weighted_score DESC
#     """
#     conn = sqlite3.connect(db_path)
#     conn.row_factory = sqlite3.Row
#     cur = conn.cursor()
#     cur.execute(sql, (report_id, report_id))
#     rows = cur.fetchall()
#     all_passes = [dict(row) for row in rows]
#     # Only keep passes with a real weighted_score
#     passes = [p for p in all_passes if p['weighted_score'] is not None]
#     if not passes:
#         conn.close()
#         return []
#
#     max_score = max(p['weighted_score'] for p in passes)
#     score_cutoff = fuzzy_threshold * max_score
#
#     # Select all passes within fuzzy threshold
#     fuzzy_passes = [p for p in passes if p['weighted_score'] >= score_cutoff]
#
#     # Always sort descending by weighted_score
#     fuzzy_passes = sorted(fuzzy_passes, key=lambda p: p['weighted_score'], reverse=True)
#
#     # Guarantee at least one pass (top scoring) if fuzzy filter yields none
#     if not fuzzy_passes:
#         fuzzy_passes = [max(passes, key=lambda p: p['weighted_score'])]
#
#     # Optionally limit to top_n (if many within fuzzy)
#     if len(fuzzy_passes) > top_n:
#         fuzzy_passes = fuzzy_passes[:top_n]
#
#     conn.close()
#     return fuzzy_passes

def get_top_n_passes(db_path, report_id, weights, top_n, fuzzy_threshold=0.9, dist_threshold=0.5):
    """
    Two-step selection:
    1. Only consider passes within a certain normalized_total_distance_to_good threshold.
    2. Rank by weighted_score, apply fuzzy threshold and top_n.
    If no pass matches the distance filter, fallback to top normalized_total_distance_to_good, limit by top_n.
    """
    import sqlite3

    sql = """
    WITH stats AS (
        SELECT
            MIN(profit) AS min_np, MAX(profit) AS max_np,
            MIN((profit/NULLIF(drawdown_abs,0))) AS min_rf, MAX((profit/NULLIF(drawdown_abs,0))) AS max_rf,
            MIN(profit_factor) AS min_pf, MAX(profit_factor) AS max_pf,
            MIN(expected_payoff) AS min_ep, MAX(expected_payoff) AS max_ep,
            MIN(total_trades) AS min_tt, MAX(total_trades) AS max_tt,
            MIN(drawdown_abs) AS min_md, MAX(drawdown_abs) AS max_md,
            MIN(min_total_recovery) AS min_total_recovery,
            MIN(min_trades) AS min_trades,
            MIN(min_max_drawdown) AS min_max_drawdown
        FROM optimization_passes
        WHERE report_id = ?
    )
    , scored AS (
        SELECT
            p.*,
            -- Calculate metrics for scoring/normalization
            (1.0 * (p.profit      - s.min_np) / NULLIF(s.max_np  - s.min_np, 0) +
             2.0 * ((p.profit/NULLIF(p.drawdown_abs,0)) - s.min_rf) / NULLIF(s.max_rf  - s.min_rf, 0) +
             1.5 * (p.profit_factor   - s.min_pf) / NULLIF(s.max_pf  - s.min_pf, 0) +
             1.0 * (p.expected_payoff - s.min_ep) / NULLIF(s.max_ep  - s.min_ep, 0) +
             1.0 * (p.total_trades    - s.min_tt) / NULLIF(s.max_tt  - s.min_tt, 0) +
             2.0 * (s.max_md - p.drawdown_abs) / NULLIF(s.max_md - s.min_md, 0)
            ) AS weighted_score,

            CASE 
                WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0
                ELSE (s.min_total_recovery - (p.profit/NULLIF(p.drawdown_abs,0))) / NULLIF(s.min_total_recovery, 0)
            END AS recovery_factor_distance_norm,

            CASE 
                WHEN p.total_trades >= s.min_trades THEN 0
                ELSE (s.min_trades - p.total_trades) / NULLIF(s.min_trades, 0)
            END AS total_trades_distance_norm,

            CASE 
                WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0
                ELSE (p.drawdown_abs - s.min_max_drawdown) / NULLIF(s.min_max_drawdown, 0)
            END AS max_drawdown_distance_norm,

            (
                (CASE WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0 ELSE 1 END) +
                (CASE WHEN p.total_trades >= s.min_trades THEN 0 ELSE 1 END) +
                (CASE WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0 ELSE 1 END) +
                CASE WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0 ELSE (s.min_total_recovery - (p.profit/NULLIF(p.drawdown_abs,0))) / NULLIF(s.min_total_recovery, 0) END +
                CASE WHEN p.total_trades >= s.min_trades THEN 0 ELSE (s.min_trades - p.total_trades) / NULLIF(s.min_trades, 0) END +
                CASE WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0 ELSE (p.drawdown_abs - s.min_max_drawdown) / NULLIF(s.min_max_drawdown, 0) END
            ) AS normalized_total_distance_to_good
        FROM optimization_passes p, stats s
        WHERE p.report_id = ?
    ),
    deduped AS (
        SELECT MIN(id) AS id
        FROM scored
        GROUP BY profit, drawdown_abs, total_trades, weighted_score
    )
    SELECT s.*
    FROM scored s
    JOIN deduped d ON s.id = d.id
    WHERE s.normalized_total_distance_to_good < ?
    ORDER BY s.weighted_score DESC
    """

    sql_top_dist = """
    WITH stats AS (
        SELECT
            MIN(profit) AS min_np, MAX(profit) AS max_np,
            MIN((profit/NULLIF(drawdown_abs,0))) AS min_rf, MAX((profit/NULLIF(drawdown_abs,0))) AS max_rf,
            MIN(profit_factor) AS min_pf, MAX(profit_factor) AS max_pf,
            MIN(expected_payoff) AS min_ep, MAX(expected_payoff) AS max_ep,
            MIN(total_trades) AS min_tt, MAX(total_trades) AS max_tt,
            MIN(drawdown_abs) AS min_md, MAX(drawdown_abs) AS max_md,
            MIN(min_total_recovery) AS min_total_recovery,
            MIN(min_trades) AS min_trades,
            MIN(min_max_drawdown) AS min_max_drawdown
        FROM optimization_passes
        WHERE report_id = ?
    )
    , scored AS (
        SELECT
            p.*,
            (1.0 * (p.profit      - s.min_np) / NULLIF(s.max_np  - s.min_np, 0) +
             2.0 * ((p.profit/NULLIF(p.drawdown_abs,0)) - s.min_rf) / NULLIF(s.max_rf  - s.min_rf, 0) +
             1.5 * (p.profit_factor   - s.min_pf) / NULLIF(s.max_pf  - s.min_pf, 0) +
             1.0 * (p.expected_payoff - s.min_ep) / NULLIF(s.max_ep  - s.min_ep, 0) +
             1.0 * (p.total_trades    - s.min_tt) / NULLIF(s.max_tt  - s.min_tt, 0) +
             2.0 * (s.max_md - p.drawdown_abs) / NULLIF(s.max_md - s.min_md, 0)
            ) AS weighted_score,

            CASE 
                WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0
                ELSE (s.min_total_recovery - (p.profit/NULLIF(p.drawdown_abs,0))) / NULLIF(s.min_total_recovery, 0)
            END AS recovery_factor_distance_norm,

            CASE 
                WHEN p.total_trades >= s.min_trades THEN 0
                ELSE (s.min_trades - p.total_trades) / NULLIF(s.min_trades, 0)
            END AS total_trades_distance_norm,

            CASE 
                WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0
                ELSE (p.drawdown_abs - s.min_max_drawdown) / NULLIF(s.min_max_drawdown, 0)
            END AS max_drawdown_distance_norm,

            (
                (CASE WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0 ELSE 1 END) +
                (CASE WHEN p.total_trades >= s.min_trades THEN 0 ELSE 1 END) +
                (CASE WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0 ELSE 1 END) +
                CASE WHEN (p.profit/NULLIF(p.drawdown_abs,0)) >= s.min_total_recovery THEN 0 ELSE (s.min_total_recovery - (p.profit/NULLIF(p.drawdown_abs,0))) / NULLIF(s.min_total_recovery, 0) END +
                CASE WHEN p.total_trades >= s.min_trades THEN 0 ELSE (s.min_trades - p.total_trades) / NULLIF(s.min_trades, 0) END +
                CASE WHEN p.drawdown_abs <= s.min_max_drawdown THEN 0 ELSE (p.drawdown_abs - s.min_max_drawdown) / NULLIF(s.min_max_drawdown, 0) END
            ) AS normalized_total_distance_to_good
        FROM optimization_passes p, stats s
        WHERE p.report_id = ?
    ),
    deduped AS (
        SELECT MIN(id) AS id
        FROM scored
        GROUP BY profit, drawdown_abs, total_trades, weighted_score
    )
    SELECT s.*
    FROM scored s
    JOIN deduped d ON s.id = d.id
    WHERE s.normalized_total_distance_to_good IS NOT NULL
    ORDER BY s.normalized_total_distance_to_good ASC, s.weighted_score DESC
    LIMIT ?
    """

    conn = sqlite3.connect(db_path)
    #conn.execute("PRAGMA key = 'Kh78784bt!'")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(sql, (report_id, report_id, dist_threshold))
    rows = cur.fetchall()
    all_passes = [dict(row) for row in rows]
    passes = [p for p in all_passes if p['weighted_score'] is not None]

    # Fallback: If no pass is close enough to requirements, select top_n by minimal normalized_total_distance_to_good
    if not passes:
        cur.execute(sql_top_dist, (report_id, report_id, top_n))
        rows = cur.fetchall()
        conn.close()
        if rows:
            return [dict(row) for row in rows]
        else:
            return []

    max_score = max(p['weighted_score'] for p in passes)
    score_cutoff = fuzzy_threshold * max_score

    # Select all passes within fuzzy threshold
    fuzzy_passes = [p for p in passes if p['weighted_score'] >= score_cutoff]
    fuzzy_passes = sorted(fuzzy_passes, key=lambda p: p['weighted_score'], reverse=True)

    if len(fuzzy_passes) > top_n:
        fuzzy_passes = fuzzy_passes[:top_n]

    conn.close()
    return fuzzy_passes

def insert_set_file_artifacts(
    db_path,
    step_id,
    artifact_type=None,
    file_path=None,
    meta_json=None,
    link_type=None,
    link_id=None
):
    """
    Insert a row into set_file_artifacts table.
    Reads file_blob from file_path if provided and file exists.
    Arguments:
        db_path (str): Path to the SQLite database file.
        step_id (int): Foreign key referencing set_file_steps(id).
        artifact_type (str or None)
        file_path (str or None)
        meta_json (str or None)
        link_type (str or None)
        link_id (int or None)
    Returns:
        int: The id of the newly inserted row.
    """
    # Read file_blob from file_path, if it exists
    file_blob = None
    if file_path and os.path.isfile(file_path):
        with open(file_path, "rb") as f:
            file_blob = f.read()

    sql = """
        INSERT INTO set_file_artifacts (
            step_id, artifact_type, file_path, meta_json, file_blob, link_type, link_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """
    params = (step_id, artifact_type, file_path, meta_json, file_blob, link_type, link_id)
    conn = sqlite3.connect(db_path)
    #conn.execute("PRAGMA key = 'Kh78784bt!'")
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()

def process_optimization_report_topn(
    html_report_path,
    db_path,
    step_id,
    config_xlsx_path
):
    # Load both criteria and weights/top_n for maximum flexibility
    criteria = read_performance_criteria_xlsx(config_xlsx_path)
    weights, top_n, fuzzy, distance = read_optimization_config(config_xlsx_path)
    report = parse_report(html_report_path)
    report_id = insert_into_db(report, db_path, step_id, criteria)
    top_passes = get_top_n_passes(db_path, report_id, weights, top_n, fuzzy_threshold=fuzzy, dist_threshold=distance)
    result = [
        {
            "optimization_pass_id": p["id"],  # <-- Add this line
            "pass_number": p["pass_number"],
            "weighted_score": p["weighted_score"],
            "profit": p["profit"],
            "drawdown_abs": p["drawdown_abs"],
            "total_trades": p["total_trades"],
            "profit_factor": p["profit_factor"],
        }
        for p in top_passes
    ]

    insert_set_file_artifacts(
        db_path=db_path,
        step_id=step_id,
        artifact_type="optimization_report",  # for .htm
        file_path=html_report_path,
        meta_json=json.dumps({}),
        link_type="optimization_reports",
        link_id=report_id
    )

    insert_set_file_artifacts(
        db_path=db_path,
        step_id=step_id,
        artifact_type="optimization_report_gif",  # for .gif
        file_path=html_report_path.replace('.htm', '.gif'),  # assuming naming convention
        meta_json=json.dumps({}),
        link_type="optimization_reports",
        link_id=report_id
    )

    return json.dumps({"top_n_passes": result}, indent=2)

def process_optimization_report(
    html_report_path,
    db_path,
    step_id,
    perf_criteria_path=None
):
    # Load criteria from Excel if path is given
    if perf_criteria_path:
        criteria = read_performance_criteria_xlsx(perf_criteria_path)
    else:
        criteria = {}

    report = parse_report(html_report_path)
    report_id = insert_into_db(report, db_path, step_id, criteria)
    # Query for best pass_number (among passes that passed criteria)
    conn = sqlite3.connect(db_path)
    #conn.execute("PRAGMA key = 'Kh78784bt!'")
    cur = conn.cursor()
    cur.execute("""
        SELECT pass_number FROM optimization_passes 
        WHERE report_id = ? AND criteria_passed = 1
        ORDER BY score DESC 
        LIMIT 1
    """, (report_id,))
    row = cur.fetchone()
    best_pass_number = row[0] if row else None
    conn.close()

    insert_set_file_artifacts(
        db_path=db_path,
        step_id=step_id,
        artifact_type="optimization_report",  # for .htm
        file_path=html_report_path,
        meta_json=json.dumps({}),
        link_type="optimization_reports",
        link_id=report_id
    )

    insert_set_file_artifacts(
        db_path=db_path,
        step_id=step_id,
        artifact_type="optimization_report_gif",  # for .gif
        file_path=html_report_path.replace('.htm', '.gif'),  # assuming naming convention
        meta_json=json.dumps({}),
        link_type="optimization_reports",
        link_id=report_id
    )

    return json.dumps({"best_pass_number": best_pass_number})

# if __name__ == "__main__":
#     parser = argparse.ArgumentParser(description="Process MT4 Optimization HTML report and store filtered results in DB.")
#     parser.add_argument("HTML_REPORT_PATH", type=str, help="Path to the MT4 HTML optimization report")
#     parser.add_argument("DB_PATH", type=str, help="Path to the SQLite database file")
#     parser.add_argument("STEP_ID", type=int, help="Step ID for the optimization report")
#     parser.add_argument("--config_xlsx", type=str, default="C:\\Users\\Philip\\Documents\\UiPath\\MT4 Backtesting Automation\\Data\\Config.xlsx", help="Path to Config.xlsx (with performance_criteria, optimization_weights, optimization_setting)")
#     parser.add_argument("--topn", action="store_true", help="Return Top N passes instead of just the best")
#     args = parser.parse_args()
#
#     if args.topn:
#         topn_result = process_optimization_report_topn(
#             args.HTML_REPORT_PATH,
#             args.DB_PATH,
#             args.STEP_ID,
#             args.config_xlsx
#         )
#         print(topn_result)
#     else:
#         best_pass_number = process_optimization_report(
#             args.HTML_REPORT_PATH,
#             args.DB_PATH,
#             args.STEP_ID,
#             perf_criteria_path=args.config_xlsx
#         )
#         print(f"Best pass_number for step_id={args.STEP_ID}: {best_pass_number}")

# New Main to output JSON with success/error/data and prepare for pyinstall packaging for integration with uipath
# if __name__ == "__main__":
#     import argparse
#     import json
#     output = {}
#     try:
#         parser = argparse.ArgumentParser(description="Process MT4 Optimization HTML report and store filtered results in DB.")
#         parser.add_argument("HTML_REPORT_PATH", type=str, help="Path to the MT4 HTML optimization report")
#         parser.add_argument("DB_PATH", type=str, help="Path to the SQLite database file")
#         parser.add_argument("STEP_ID", type=int, help="Step ID for the optimization report")
#         parser.add_argument("--config_xlsx", type=str, default="C:\\Users\\Philip\\Documents\\UiPath\\MT4 Backtesting Automation\\Data\\Config.xlsx", help="Path to Config.xlsx (with performance_criteria, optimization_weights, optimization_setting)")
#         parser.add_argument("--topn", action="store_true", help="Return Top N passes instead of just the best")
#         args = parser.parse_args()
#
#         output["success"] = True
#         output["error"] = ""
#         if args.topn:
#             topn_result = process_optimization_report_topn(
#                 args.HTML_REPORT_PATH,
#                 args.DB_PATH,
#                 args.STEP_ID,
#                 args.config_xlsx
#             )
#             try:
#                 result_dict = json.loads(topn_result)
#                 if isinstance(result_dict, dict):
#                     output.update(result_dict)  # Flatten keys into output
#                 else:
#                     output["result"] = result_dict
#             except Exception:
#                 output["result"] = topn_result
#         else:
#             best_pass_number = process_optimization_report(
#                 args.HTML_REPORT_PATH,
#                 args.DB_PATH,
#                 args.STEP_ID,
#                 perf_criteria_path=args.config_xlsx
#             )
#             try:
#                 result_dict = json.loads(best_pass_number)
#                 if isinstance(result_dict, dict):
#                     output.update(result_dict)
#                 else:
#                     output["best_pass_number"] = result_dict
#             except Exception:
#                 output["best_pass_number"] = best_pass_number
#     except Exception as e:
#         output["success"] = False
#         output["error"] = str(e)
#     print(json.dumps(output))

#remove the argparse parser and use direct sys.argv index-based argument parsing, so your script will accept arguments in a strict positional order, making it compatible with PowerShell's --% operator (which simply passes all arguments as-is to the EXE).
if __name__ == "__main__":
    import sys
    import json

    output = {}
    try:
        # Remove '--%' from sys.argv if present (PowerShell --% operator)
        if '--%' in sys.argv:
            sys.argv.remove('--%')

        # Usage: HTML_REPORT_PATH DB_PATH STEP_ID [CONFIG_XLSX] [TOPN]
        if len(sys.argv) < 4:
            print(json.dumps({
                "success": False,
                "error": "Insufficient arguments. Usage: HTML_REPORT_PATH DB_PATH STEP_ID [CONFIG_XLSX] [TOPN]"
            }))
            sys.exit(1)

        HTML_REPORT_PATH = sys.argv[1]
        DB_PATH = sys.argv[2]
        STEP_ID = int(sys.argv[3])

        # Defaults
        CONFIG_XLSX = "C:\\Users\\Philip\\Documents\\UiPath\\MT4 Backtesting Automation\\Data\\Config.xlsx"
        TOPN = False

        if len(sys.argv) > 4:
            CONFIG_XLSX = sys.argv[4]
        if len(sys.argv) > 5:
            # Acceptable values for TOPN: "true", "True", "1"
            TOPN = str(sys.argv[5]).lower() in ['true', '1']

        output["success"] = True
        output["error"] = ""

        if TOPN:
            topn_result = process_optimization_report_topn(
                HTML_REPORT_PATH,
                DB_PATH,
                STEP_ID,
                CONFIG_XLSX
            )
            try:
                result_dict = json.loads(topn_result)
                if isinstance(result_dict, dict):
                    output.update(result_dict)
                else:
                    output["result"] = result_dict
            except Exception:
                output["result"] = topn_result
        else:
            best_pass_number = process_optimization_report(
                HTML_REPORT_PATH,
                DB_PATH,
                STEP_ID,
                perf_criteria_path=CONFIG_XLSX
            )
            try:
                result_dict = json.loads(best_pass_number)
                if isinstance(result_dict, dict):
                    output.update(result_dict)
                else:
                    output["best_pass_number"] = result_dict
            except Exception:
                output["best_pass_number"] = best_pass_number
    except Exception as e:
        output["success"] = False
        output["error"] = str(e)
    print(json.dumps(output))
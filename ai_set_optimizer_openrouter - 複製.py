import argparse
import json
import os
import sqlite3
import re
import requests
from collections import Counter, defaultdict

from set_file_updater import update_parameters

def read_file_head(path, max_chars=8000):
    try:
        with open(path, encoding='utf-8') as f:
            return f.read(max_chars)
    except Exception:
        return ""

def load_prompt_template(template_path):
    with open(template_path, encoding="utf-8") as f:
        return f.read()

def build_ancestry_chain(conn, start_set_file_name):
    cur = conn.cursor()
    cur.execute("SELECT set_file_name, input_set_file FROM test_metrics")
    file_map = {row[0]: row[1] for row in cur.fetchall()}
    ancestry = []
    current = start_set_file_name
    while current and current not in ancestry:
        ancestry.append(current)
        input_file = file_map.get(current)
        if input_file:
            parent = os.path.basename(input_file)
            if parent == current:
                break
            current = parent
        else:
            break
    return ancestry

def parse_param_to_section(spec_content):
    param_to_section = {}
    lines = [l.strip() for l in spec_content.splitlines() if l.strip()]
    if not lines:
        return param_to_section
    header = [x.strip() for x in lines[0].split(",")]
    try:
        section_idx = header.index("Section")
        param_idx = header.index("Parameter")
    except ValueError:
        section_idx = 0
        param_idx = 1
    for line in lines[1:]:
        parts = [x.strip() for x in re.split(r',(?=(?:[^\"]*\"[^\"]*\")*[^\"]*$)', line)]
        if len(parts) > max(section_idx, param_idx):
            section = parts[section_idx]
            pname = parts[param_idx]
            if pname:
                param_to_section[pname] = section
    return param_to_section

def fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section=None):
    if not ancestry:
        return []
    placeholders = ','.join(['?'] * len(ancestry))
    sql = f"""
    SELECT p.parameter_name
    FROM set_file_steps AS s
    JOIN optimization_suggestion AS sug ON s.id = sug.step_id
    JOIN optimization_parameter AS p ON sug.id = p.suggestion_id
    JOIN test_metrics AS tm ON tm.step_id = s.id
    WHERE tm.set_file_name IN ({placeholders})
    """
    cur = conn.cursor()
    cur.execute(sql, ancestry)
    rows = [row[0] for row in cur.fetchall()]
    if param_to_section is not None:
        return [(param_to_section.get(p, "Unknown Section"), p) for p in rows]
    else:
        return [("Unknown Section", p) for p in rows]

def make_suggestion_history_summary_block(conn, set_file_name, spec_content=None):
    param_to_section = None
    if spec_content:
        param_to_section = parse_param_to_section(spec_content)
    ancestry = build_ancestry_chain(conn, set_file_name)
    section_param_pairs = fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section)
    section_param_counter = defaultdict(Counter)
    for section, pname in section_param_pairs:
        section_param_counter[section][pname] += 1
    lines = []
    for section in sorted(section_param_counter):
        lines.append(f"Section: {section}")
        for pname, count in sorted(section_param_counter[section].items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {pname}: {count}   # times previously suggested")
        lines.append("")  # Blank line for separation
    return "\n".join(lines).strip()

def build_prompt_with_history(
    prompt_template,
    base_parameters,
    set_content,
    phoenix_spec_content,
    summary_csv_content,
    performance_metrics_block,
    suggestion_history_summary_block
):
    prompt = prompt_template.format(
        base_parameters=", ".join(base_parameters),
        performance_metrics_block=performance_metrics_block,
        suggestion_history_summary_block=suggestion_history_summary_block
    )
    prompt = prompt.replace("[Upload .set FILE or paste content here]", set_content)
    prompt = prompt.replace("[Upload PhoenixSpec.csv or paste content here]", phoenix_spec_content)
    prompt = prompt.replace("[Upload SUMMARY.csv or paste content here]", summary_csv_content)
    return prompt

def extract_json_objects_from_response(text):
    code_blocks = re.findall(r"```json(.*?)```", text, re.DOTALL)
    results = []
    for block in code_blocks:
        block = block.strip()
        try:
            obj = json.loads(block)
            results.append(obj)
        except Exception:
            continue
    return results

def save_optimization_suggestion_to_db(
    db_path, step_id, mode_sections_obj, param_array
):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute(
        "INSERT INTO optimization_suggestion (step_id, mode) VALUES (?, ?)",
        (step_id, mode_sections_obj['mode'])
    )
    suggestion_id = c.lastrowid
    for section in mode_sections_obj['sections']:
        c.execute(
            "INSERT INTO optimization_section (suggestion_id, section_name, explanation) VALUES (?, ?, ?)",
            (suggestion_id, section['name'], section['explanation'])
        )
    for param in param_array:
        c.execute(
            "INSERT INTO optimization_parameter (suggestion_id, parameter_name, start, end, step, reason) VALUES (?, ?, ?, ?, ?, ?)",
            (
                suggestion_id,
                param['name'],
                param.get('start'),
                param.get('end'),
                param.get('step'),
                param.get('reason')
            )
        )
    conn.commit()
    conn.close()
    return suggestion_id

import openpyxl

def get_performance_metrics_block(config_xlsx_path, sheet_name="performance_criteria"):
    wb = openpyxl.load_workbook(config_xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    lines = []
    for row in rows:
        if row and row[0]:
            key = str(row[0])
            value = str(row[1]) if len(row) > 1 and row[1] is not None else ""
            explanation = str(row[2]) if len(row) > 2 and row[2] is not None else ""
            lines.append(f"{key}: {value}   # {explanation}")
    return "\n".join(lines)

def call_openrouter(prompt, model, api_key):
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are an expert in MetaTrader 4/5 optimization and parameter tuning."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.1,
        "max_tokens": 1600
    }
    response = requests.post(url, headers=headers, json=data)
    response.raise_for_status()
    result = response.json()
    return result["choices"][0]["message"]["content"]

def coverage_voting(all_param_arrays):
    # all_param_arrays: list of lists (each list from a model)
    param_ranges = defaultdict(list)
    for param_array in all_param_arrays:
        for s in param_array:
            param_ranges[s["name"]].append((s.get("start"), s.get("end"), s.get("step")))
    merged = []
    for name, ranges in param_ranges.items():
        numeric_ranges = [r for r in ranges if r[0] is not None and r[1] is not None and r[2] is not None]
        if not numeric_ranges:
            continue
        try:
            min_start = min(r[0] for r in numeric_ranges)
            max_end = max(r[1] for r in numeric_ranges)
            min_step = min(r[2] for r in numeric_ranges)
        except Exception:
            # fallback: skip if any value is not comparable
            continue
        merged.append({
            "name": name,
            "start": min_start,
            "end": max_end,
            "step": min_step,
            "reason": f"Coverage of all model suggestions: {[(r[0], r[1]) for r in ranges]}"
        })
    return merged

def suggest_mode_and_sections_and_params_openrouter(
    template_path,
    base_parameters,
    set_path,
    spec_path,
    summary_path,
    openrouter_api_key,
    output_path,
    db_path,
    step_id,
    config_xlsx_path=None,
    suggestion_json_path=None,
    models=None,
):
    prompt_template = load_prompt_template(template_path)
    set_content = read_file_head(set_path, 8000)
    phoenix_spec_content = read_file_head(spec_path, 8000)
    summary_csv_content = read_file_head(summary_path, 8000)
    base_parameters_list = [s.strip() for s in base_parameters.split(",") if s.strip()]
    if config_xlsx_path:
        performance_metrics_block = get_performance_metrics_block(config_xlsx_path)
    else:
        performance_metrics_block = ""
    set_file_name = os.path.basename(set_path)
    conn = sqlite3.connect(db_path)
    suggestion_history_summary_block = make_suggestion_history_summary_block(
        conn, set_file_name, spec_content=phoenix_spec_content
    )
    conn.close()
    prompt = build_prompt_with_history(
        prompt_template,
        base_parameters_list,
        set_content,
        phoenix_spec_content,
        summary_csv_content,
        performance_metrics_block,
        suggestion_history_summary_block
    )

    # List of models to query
    if not models:
        models = ["openai/gpt-4o"]  # default to GPT-4o

    print(f"Models: {models}")
    print(f"Prompt size: {len(prompt)} chars")
    try:
        import tiktoken
        enc = tiktoken.encoding_for_model("gpt-4o")
        token_count = len(enc.encode(prompt))
        print(f"Prompt token count: {token_count}")
    except Exception as e:
        print(f"Warning: Could not count tokens ({e}).")

    all_mode_sections = []
    all_param_arrays = []
    for model in models:
        print(f"Querying OpenRouter model: {model}")
        try:
            resp_content = call_openrouter(prompt, model, openrouter_api_key)
        except Exception as e:
            print(f"[ERROR] OpenRouter call failed for model {model}: {e}")
            continue
        objs = extract_json_objects_from_response(resp_content)
        mode_sections_obj = None
        param_array = None
        for obj in objs:
            if isinstance(obj, dict) and "mode" in obj and "sections" in obj:
                mode_sections_obj = obj
            elif isinstance(obj, list):
                param_array = obj
        if not mode_sections_obj or not param_array:
            suggestions_path = output_path + f".{model.replace('/','_')}.suggestions.txt"
            with open(suggestions_path, "w", encoding="utf-8") as f:
                f.write(resp_content)
            print(f"[ERROR] No valid suggestion JSON found from {model}. Raw response written to: {suggestions_path}")
            continue
        all_mode_sections.append(mode_sections_obj)
        all_param_arrays.append(param_array)

    if not all_param_arrays:
        print("[ERROR] No valid parameter suggestions from any model.")
        return None

    # Merge parameter arrays by coverage
    merged_param_array = coverage_voting(all_param_arrays)
    # Use the first mode/sections block as representative (can be improved with voting/merging)
    final_mode_sections_obj = all_mode_sections[0] if all_mode_sections else {"mode": "", "sections": []}

    # Save to DB
    suggestion_id = save_optimization_suggestion_to_db(
        db_path, step_id, final_mode_sections_obj, merged_param_array
    )

    # Optionally write to files for backup/debugging
    if suggestion_json_path:
        with open(suggestion_json_path, "w", encoding="utf-8") as f:
            json.dump({
                "mode_sections": final_mode_sections_obj,
                "parameters": merged_param_array
            }, f, indent=2)
    else:
        with open(output_path + ".suggestions.json", "w", encoding="utf-8") as f:
            json.dump({
                "mode_sections": final_mode_sections_obj,
                "parameters": merged_param_array
            }, f, indent=2)

    # Update parameters in .set file using only parameter array
    update_parameters(set_path, merged_param_array, output_path)
    print(f"DB: suggestion_id={suggestion_id}. Updated .set file written to: {output_path}")

    return output_path

def main():
    parser = argparse.ArgumentParser(description="AI-powered .set file optimizer using OpenRouter and multi-model parameter suggestion coverage voting.")
    parser.add_argument("--template", required=True, help="Path to ai_parameter_suggestion_prompt_template file (Markdown)")
    parser.add_argument("--base-parameters", required=True, help="Comma-separated base parameters to highlight for optimization")
    parser.add_argument("--set", required=True, help="Input .set file path")
    parser.add_argument("--spec", required=True, help="PhoenixSpec.csv file path")
    parser.add_argument("--summary", required=True, help="SUMMARY.csv/backtest summary file path")
    parser.add_argument("--openrouter-api-key", required=True, help="OpenRouter API key")
    parser.add_argument("--output", required=True, help="Output .set file path")
    parser.add_argument("--db-path", required=True, help="Path to the SQLite database")
    parser.add_argument("--step-id", required=True, type=int, help="set_file_steps.id to link suggestion to")
    parser.add_argument("--config_xlsx_path", required=False, help="Config.xlsx for performance metrics (optional, for performance criteria)")
    parser.add_argument("--suggestion-json", required=False, help="Output path for suggestion JSON file (default: output.suggestions.json)")
    parser.add_argument("--models", required=False, help="Comma-separated list of OpenRouter model names (e.g. openai/gpt-4o,anthropic/claude-3-opus)")
    args = parser.parse_args()

    models = None
    if args.models:
        models = [m.strip() for m in args.models.split(",") if m.strip()]

    output_path = suggest_mode_and_sections_and_params_openrouter(
        template_path=args.template,
        base_parameters=args.base_parameters,
        set_path=args.set,
        spec_path=args.spec,
        summary_path=args.summary,
        openrouter_api_key=args.openrouter_api_key,
        output_path=args.output,
        db_path=args.db_path,
        step_id=args.step_id,
        config_xlsx_path=args.config_xlsx_path,
        suggestion_json_path=args.suggestion_json,
        models=models,
    )

    if output_path:
        print(output_path)

if __name__ == "__main__":
    main()
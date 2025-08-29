import argparse
import json
import openai
import os
import sqlite3
import re
from collections import Counter, defaultdict

from set_file_updater import update_parameters

def read_file_head(path, max_chars=8000):
    try:
        with open(path, encoding='utf-8') as f:
            return f.read(max_chars)
    except Exception:
        return ""

def load_prompt_template(template_path):
    with open(template_path, encoding="utf-8") as f:
        return f.read()

# --- SUGGESTION HISTORY BLOCK INTEGRATION START ---

def build_ancestry_chain(conn, start_set_file_name):
    """Recursively find all ancestor set_file_names for suggestion history."""
    cur = conn.cursor()
    cur.execute("SELECT set_file_name, input_set_file FROM test_metrics")
    file_map = {row[0]: row[1] for row in cur.fetchall()}

    ancestry = []
    current = start_set_file_name
    while current and current not in ancestry:
        ancestry.append(current)
        input_file = file_map.get(current)
        if input_file:
            parent = os.path.basename(input_file)
            if parent == current:
                break
            current = parent
        else:
            break
    return ancestry

def parse_param_to_section(spec_content):
    """
    Parse the parameter spec CSV content (string) and build a dict: {param_name: section_name}
    """
    param_to_section = {}
    lines = [l.strip() for l in spec_content.splitlines() if l.strip()]
    if not lines:
        return param_to_section
    header = [x.strip() for x in lines[0].split(",")]
    try:
        section_idx = header.index("Section")
        param_idx = header.index("Parameter")
    except ValueError:
        # Fallback: try to find columns by similar names
        section_idx = 0
        param_idx = 1
    for line in lines[1:]:
        parts = [x.strip() for x in re.split(r',(?=(?:[^\"]*\"[^\"]*\")*[^\"]*$)', line)]
        if len(parts) > max(section_idx, param_idx):
            section = parts[section_idx]
            pname = parts[param_idx]
            if pname:
                param_to_section[pname] = section
    return param_to_section

def fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section=None):
    """
    Fetch all parameter suggestions for the ancestry, with section info (if available).
    Returns a list of (section, parameter) tuples.
    """
    if not ancestry:
        return []
    placeholders = ','.join(['?'] * len(ancestry))
    sql = f"""
    SELECT p.parameter_name
    FROM set_file_steps AS s
    JOIN optimization_suggestion AS sug ON s.id = sug.step_id
    JOIN optimization_parameter AS p ON sug.id = p.suggestion_id
    JOIN test_metrics AS tm ON tm.step_id = s.id
    WHERE tm.set_file_name IN ({placeholders})
    """
    cur = conn.cursor()
    cur.execute(sql, ancestry)
    rows = [row[0] for row in cur.fetchall()]
    # Map parameter to section if possible
    if param_to_section is not None:
        return [(param_to_section.get(p, "Unknown Section"), p) for p in rows]
    else:
        return [("Unknown Section", p) for p in rows]

def make_suggestion_history_summary_block(conn, set_file_name, spec_content=None):
    # Build param_to_section if possible
    param_to_section = None
    if spec_content:
        param_to_section = parse_param_to_section(spec_content)
    ancestry = build_ancestry_chain(conn, set_file_name)
    section_param_pairs = fetch_suggestions_for_ancestry_sections(conn, ancestry, param_to_section)
    # Count suggestions per section/parameter
    section_param_counter = defaultdict(Counter)
    for section, pname in section_param_pairs:
        section_param_counter[section][pname] += 1

    # Format as requested
    lines = []
    for section in sorted(section_param_counter):
        lines.append(f"Section: {section}")
        for pname, count in sorted(section_param_counter[section].items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {pname}: {count}   # times previously suggested")
        lines.append("")  # Blank line for separation

    return "\n".join(lines).strip()

def build_prompt_with_history(
    prompt_template,
    base_parameters,
    set_content,
    phoenix_spec_content,
    summary_csv_content,
    performance_metrics_block,
    suggestion_history_summary_block
):

    prompt = prompt_template.format(
        base_parameters=", ".join(base_parameters),
        performance_metrics_block=performance_metrics_block,
        suggestion_history_summary_block=suggestion_history_summary_block
    )

    # Fill in set, spec, and summary
    prompt = prompt.replace("[Upload .set FILE or paste content here]", set_content)
    prompt = prompt.replace("[Upload PhoenixSpec.csv or paste content here]", phoenix_spec_content)
    prompt = prompt.replace("[Upload SUMMARY.csv or paste content here]", summary_csv_content)
    return prompt

# --- SUGGESTION HISTORY BLOCK INTEGRATION END ---

def extract_json_objects_from_response(text):
    """
    Extracts all JSON objects/arrays from ```json ... ``` code blocks.
    Returns a list of loaded objects.
    """
    code_blocks = re.findall(r"```json(.*?)```", text, re.DOTALL)
    results = []
    for block in code_blocks:
        block = block.strip()
        try:
            obj = json.loads(block)
            results.append(obj)
        except Exception:
            continue
    return results

def save_optimization_suggestion_to_db(
    db_path, step_id, mode_sections_obj, param_array
):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    # Insert main suggestion
    c.execute(
        "INSERT INTO optimization_suggestion (step_id, mode) VALUES (?, ?)",
        (step_id, mode_sections_obj['mode'])
    )
    suggestion_id = c.lastrowid
    # Insert sections
    for section in mode_sections_obj['sections']:
        c.execute(
            "INSERT INTO optimization_section (suggestion_id, section_name, explanation) VALUES (?, ?, ?)",
            (suggestion_id, section['name'], section['explanation'])
        )
    # Insert parameters
    for param in param_array:
        c.execute(
            "INSERT INTO optimization_parameter (suggestion_id, parameter_name, start, end, step, reason) VALUES (?, ?, ?, ?, ?, ?)",
            (
                suggestion_id,
                param['name'],
                param.get('start'),
                param.get('end'),
                param.get('step'),
                param.get('reason')
            )
        )
    conn.commit()
    conn.close()
    return suggestion_id

import openpyxl

def get_performance_metrics_block(config_xlsx_path, sheet_name="performance_criteria"):
    wb = openpyxl.load_workbook(config_xlsx_path, data_only=True)
    ws = wb[sheet_name]
    # Get headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    lines = []
    for row in rows:
        if row and row[0]:
            # Defensive: handle missing columns
            key = str(row[0])
            value = str(row[1]) if len(row) > 1 and row[1] is not None else ""
            explanation = str(row[2]) if len(row) > 2 and row[2] is not None else ""
            lines.append(f"{key}: {value}   # {explanation}")
    return "\n".join(lines)

def suggest_mode_and_sections_and_params(
    template_path,
    base_parameters,
    set_path,
    spec_path,
    summary_path,
    api_key,
    output_path,
    db_path,
    step_id,
    config_xlsx_path=None,
    suggestion_json_path=None
):
    # Load contents
    prompt_template = load_prompt_template(template_path)
    set_content = read_file_head(set_path, 8000)
    phoenix_spec_content = read_file_head(spec_path, 8000)
    summary_csv_content = read_file_head(summary_path, 8000)

    # Handle multiple base parameters
    base_parameters_list = [s.strip() for s in base_parameters.split(",") if s.strip()]

    if config_xlsx_path:
        performance_metrics_block = get_performance_metrics_block(config_xlsx_path)
    else:
        performance_metrics_block = ""

    # --- Get set_file_name for the ancestry chain (assume .set file name) ---
    set_file_name = os.path.basename(set_path)

    # --- Open DB connection for ancestry/suggestion history ---
    conn = sqlite3.connect(db_path)
    suggestion_history_summary_block = make_suggestion_history_summary_block(
        conn, set_file_name, spec_content=phoenix_spec_content
    )
    conn.close()

    # --- Build prompt (integrated with suggestion_history_summary_block) ---
    prompt = build_prompt_with_history(
        prompt_template,
        base_parameters_list,
        set_content,
        phoenix_spec_content,
        summary_csv_content,
        performance_metrics_block,
        suggestion_history_summary_block
    )

    # print("--- PROMPT ---")
    # print(prompt)
    # print("--- END ---")

    # Print token count for verification
    try:
        import tiktoken
        enc = tiktoken.encoding_for_model("gpt-4o")
        token_count = len(enc.encode(prompt))
        print(f"Prompt token count: {token_count}")
    except Exception as e:
        print(f"Warning: Could not count tokens ({e}).")

    # OpenAI call
    client = openai.OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are an expert in MetaTrader 4/5 optimization and parameter tuning."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.1,
        max_tokens=1600,
    )
    resp_content = response.choices[0].message.content

    # Extract both JSON outputs
    objs = extract_json_objects_from_response(resp_content)
    # Find the mode/sections object and parameter array
    mode_sections_obj = None
    param_array = None
    for obj in objs:
        if isinstance(obj, dict) and "mode" in obj and "sections" in obj:
            mode_sections_obj = obj
        elif isinstance(obj, list):
            param_array = obj
    if not mode_sections_obj or not param_array:
        # Write raw response for debugging
        suggestions_path = output_path + ".suggestions.txt"
        with open(suggestions_path, "w", encoding="utf-8") as f:
            f.write(resp_content)
        print(f"[ERROR] No valid suggestion JSON found. Raw response written to: {suggestions_path}")
        return None

    # Save to DB
    suggestion_id = save_optimization_suggestion_to_db(
        db_path, step_id, mode_sections_obj, param_array
    )

    # Optionally write to files for backup/debugging
    if suggestion_json_path:
        with open(suggestion_json_path, "w", encoding="utf-8") as f:
            json.dump({
                "mode_sections": mode_sections_obj,
                "parameters": param_array
            }, f, indent=2)
    else:
        with open(output_path + ".suggestions.json", "w", encoding="utf-8") as f:
            json.dump({
                "mode_sections": mode_sections_obj,
                "parameters": param_array
            }, f, indent=2)

    # Update parameters in .set file using only parameter array
    update_parameters(set_path, param_array, output_path)
    print(f"DB: suggestion_id={suggestion_id}. Updated .set file written to: {output_path}")

    return output_path

# Example CLI usage for the new function
def main():
    parser = argparse.ArgumentParser(description="AI-powered .set file optimizer using OpenAI and parameter suggestion prompt template.")
    parser.add_argument("--template", required=True, help="Path to ai_parameter_suggestion_prompt_template file (Markdown)")
    parser.add_argument("--base-parameters", required=True, help="Comma-separated base parameters to highlight for optimization")
    parser.add_argument("--set", required=True, help="Input .set file path")
    parser.add_argument("--spec", required=True, help="PhoenixSpec.csv file path")
    parser.add_argument("--summary", required=True, help="SUMMARY.csv/backtest summary file path")
    parser.add_argument("--api-key", required=True, help="OpenAI API key")
    parser.add_argument("--output", required=True, help="Output .set file path")
    parser.add_argument("--db-path", required=True, help="Path to the SQLite database")
    parser.add_argument("--step-id", required=True, type=int, help="set_file_steps.id to link suggestion to")
    parser.add_argument("--config_xlsx_path", required=False, help="Config.xlsx for performance metrics (optional, for performance criteria)")
    parser.add_argument("--suggestion-json", required=False, help="Output path for suggestion JSON file (default: output.suggestions.json)")
    args = parser.parse_args()

    output_path = suggest_mode_and_sections_and_params(
        template_path=args.template,
        base_parameters=args.base_parameters,
        set_path=args.set,
        spec_path=args.spec,
        summary_path=args.summary,
        api_key=args.api_key,
        output_path=args.output,
        db_path=args.db_path,
        step_id=args.step_id,
        config_xlsx_path=args.config_xlsx_path,
        suggestion_json_path=args.suggestion_json
    )

    if output_path:
        print(output_path)

if __name__ == "__main__":
    main()